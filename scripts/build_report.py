"""Build report (Excel + HTML) from cached timelines + extended fields."""
from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OPEN_STATUSES = {
    "Abierto", "Creado", "Devuelto a cliente", "En preparación presupuesto",
    "En cola taller", "En préstamo", "En reparación", "Enviado a técnico externo",
    "Equipo devuelto", "Equipo enviado", "Esperando inicio reparación",
    "Esperando respuesta cliente a presupuesto", "Formulario en curso",
    "Formulario enviado a calidad", "Gestionado transporte", "Inspección de salida",
    "Investigación", "Material enviado", "Pendiente agendar llamada",
    "Pendiente asignar técnico", "Pendiente confirmación presupuesto",
    "Pendiente definir servicio externo", "Pendiente recogida",
    "Presupuesto aceptado", "Presupuesto enviado",
    "Presupuesto preparado pendiente de enviar", "Queja creada",
    "Recepcionado SAT", "Reportado", "Solicitado",
}

FUNNEL_ORDER = [
    "Abierto", "Recepcionado SAT", "Pendiente recogida", "Gestionado transporte",
    "Pendiente asignar técnico", "En cola taller", "En preparación presupuesto",
    "Presupuesto preparado pendiente de enviar", "Pendiente confirmación presupuesto",
    "Esperando inicio reparación", "En reparación", "Inspección de salida",
    "Devuelto a cliente", "Pendiente agendar llamada", "Pendiente definir servicio externo",
    "Esperando respuesta cliente a presupuesto", "Enviado a técnico externo",
]

FUNNEL_TAG = {
    "Abierto": "Inicio", "Recepcionado SAT": "Interna", "Pendiente recogida": "Interna",
    "Gestionado transporte": "Interna", "Pendiente asignar técnico": "Interna",
    "En cola taller": "Interna", "En preparación presupuesto": "Interna",
    "Presupuesto preparado pendiente de enviar": "Interna",
    "Pendiente confirmación presupuesto": "Interna",
    "Esperando inicio reparación": "Interna", "En reparación": "Interna",
    "Inspección de salida": "Interna", "Devuelto a cliente": "Interna",
    "Pendiente agendar llamada": "Online",
    "Pendiente definir servicio externo": "Externa",
    "Esperando respuesta cliente a presupuesto": "Externa",
    "Enviado a técnico externo": "Externa",
}

CHAIN_ORDER = ["Elha", "Sin Vello", "Dermasana", "Smart Duck",
               "Epil Point", "Laser Factory", "Unico Italia", "Otros"]

FUNNEL_COLOR = {"Inicio": "#9aa5b1", "Interna": "#1f6feb", "Online": "#cb6f0a", "Externa": "#a23b72"}
COLOR_GREEN = "#1f8a4c"
COLOR_YELLOW = "#b8860b"
COLOR_RED = "#c0392b"


def format_int_dot(n):
    """Formatea un número con separador de miles (1.234.567). '' si vacío."""
    if n == '' or n is None:
        return ''
    try:
        v = int(n)
    except (TypeError, ValueError):
        try:
            v = int(float(n))
        except Exception:
            return str(n)
    return f"{v:,}".replace(",", ".")


def chain_of(cliente, loc):
    c = (cliente or "").lower()
    l = (loc or "").lower()
    if "elha" in c: return "Elha"
    if "sin vello" in c or "sinvello" in c: return "Sin Vello"
    if "dermasana" in c: return "Dermasana"
    if "smart duck" in c or "smartduck" in c: return "Smart Duck"
    if "epil point" in c or "epilpoint" in c: return "Epil Point"
    if "laser factory" in c or "laserfactory" in c or "laser fcatory" in c: return "Laser Factory"
    if any(t in c for t in ("centro unico", "centri unico", "centros unico", "beauty cool")):
        ital = ("italia", "italy", "roma", "milano", "napoli", "torino", "bologna",
                "firenze", "ostia", "orio", "giugliano", "andria", "castellammare",
                "piacenza", "novara", "aprilia", "euroma")
        if any(ct in l for ct in ital):
            return "Unico Italia"
        return "Otros"
    return "Otros"


def parse_iso(s):
    if not s:
        return None
    s2 = s.replace("Z", "+00:00")
    s2 = re.sub(r"([+\-]\d{2})(\d{2})$", r"\1:\2", s2)
    try:
        return datetime.fromisoformat(s2)
    except ValueError:
        return None


def status_at(ticket, cutoff_dt):
    created = parse_iso(ticket.get("created"))
    if not created or created > cutoff_dt:
        return None
    transitions = ticket.get("transitions", [])
    if not transitions:
        return ticket.get("current_status")
    current = transitions[0][1]
    for ts, from_s, to_s in transitions:
        when = parse_iso(ts)
        if when and when <= cutoff_dt:
            current = to_s
        else:
            break
    return current


def last_status_change_dt(ticket):
    transitions = ticket.get("transitions", [])
    if transitions:
        last = transitions[-1]
        d = parse_iso(last[0])
        if d:
            return d
    return parse_iso(ticket.get("created"))


def is_open(s):
    return s in OPEN_STATUSES


def compute_cutoffs(today):
    """Devuelve hasta 21 cortes únicos: 6 fin-de-mes + 15 días laborables.
    Si un fin-de-mes cae en uno de los 15 días laborables, se conserva una sola vez."""
    cutoffs = []
    seen = set()
    y, m = today.year, today.month
    monthly = []
    for _ in range(6):
        m -= 1
        if m == 0:
            m, y = 12, y - 1
        nm, ny = (m + 1, y) if m < 12 else (1, y + 1)
        monthly.append(date(ny, nm, 1) - timedelta(days=1))
    monthly.sort()
    for d in monthly:
        lbl = d.strftime("%Y-%m-%d")
        if lbl in seen:
            continue
        seen.add(lbl)
        cutoffs.append((lbl, d))
    bd, d = [], today
    while len(bd) < 15:
        if d.weekday() < 5:
            bd.append(d)
        d -= timedelta(days=1)
    bd.sort()
    for x in bd:
        lbl = x.strftime("%Y-%m-%d")
        if lbl in seen:
            continue
        seen.add(lbl)
        cutoffs.append((lbl, x))
    return cutoffs


def replay_opens(cache, cutoffs):
    opens = {lbl: [] for lbl, _ in cutoffs}
    for key, t in cache.get("tickets", {}).items():
        chain = chain_of(t.get("cliente", ""), t.get("loc", ""))
        for lbl, d in cutoffs:
            cdt = datetime(d.year, d.month, d.day, 23, 59, 59, tzinfo=timezone.utc)
            s = status_at(t, cdt)
            if is_open(s):
                opens[lbl].append((key, chain, s))
    return opens


def fmt_short(lbl):
    try:
        dt = datetime.fromisoformat(lbl)
        months = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
        return f"{dt.day:02d} {months[dt.month-1]}"
    except Exception:
        return lbl


def days_since(dt):
    if not dt:
        return None
    now = datetime.now(timezone.utc)
    return (now - dt).days


def days_bucket(d):
    if d is None:
        return "y"
    if d <= 5:
        return "g"
    if d <= 15:
        return "y"
    return "r"


def is_created_today(ticket):
    """True si el ticket fue creado hoy (UTC)."""
    created = parse_iso(ticket.get("created"))
    if not created:
        return False
    today = datetime.now(timezone.utc).date()
    return created.date() == today


def compute_status_changes_today(cache):
    """Para cada (estado, cadena) cuenta tickets que ENTRARON hoy (>= 00:00 UTC hoy)
    y SALIERON hoy. Considera tickets abiertos hoy o que estaban abiertos ayer.
    Returns: dict {(estado, cadena): {'in': int, 'out': int}}
    """
    today = datetime.now(timezone.utc).date()
    cutoff_ayer = datetime(today.year, today.month, today.day, 0, 0, 0, tzinfo=timezone.utc) - timedelta(seconds=1)
    changes = {}
    for key, t in cache.get("tickets", {}).items():
        chain = chain_of(t.get("cliente", ""), t.get("loc", ""))
        status_ayer = status_at(t, cutoff_ayer)
        status_hoy = t.get("current_status")
        if status_ayer == status_hoy:
            continue
        if status_ayer and is_open(status_ayer):
            k_out = (status_ayer, chain)
            changes.setdefault(k_out, {"in": 0, "out": 0})["out"] += 1
        if status_hoy and is_open(status_hoy):
            k_in = (status_hoy, chain)
            changes.setdefault(k_in, {"in": 0, "out": 0})["in"] += 1
    return changes


# Estados donde la "persona" relevante es el assignee (Persona Asignada de Jira).
# Desde el siguiente estado en adelante, se usa "Técnico taller" (customfield_10247).
ASIGNADO_STATES = {
    "Abierto", "Recepcionado SAT", "Pendiente recogida", "Gestionado transporte",
    "Pendiente asignar técnico", "En cola taller", "En preparación presupuesto",
    "Presupuesto preparado pendiente de enviar", "Pendiente confirmación presupuesto",
    "Esperando inicio reparación",
}

# Estados donde "de verdad estamos esperando al técnico" (suma destacada en la tabla).
# Solo cuando el técnico realmente está trabajando o tiene que empezar (no incluye colas
# o esperas externas como cliente / transporte / inspección).
ESPERANDO_TECNICO_STATES = {
    "En preparación presupuesto",
    "Esperando inicio reparación",
    "En reparación",
}


def person_for_status(ticket, status):
    """Devuelve la persona relevante para un (ticket, estado).
    Hasta 'Esperando inicio reparación' (incluido) → assignee (Persona Asignada).
    Desde 'En reparación' en adelante → tec_taller."""
    if status in ASIGNADO_STATES:
        return (ticket.get("asignado") or "").strip() or "(Sin asignar)"
    return (ticket.get("tec_taller") or "").strip() or "(Sin asignar)"


def compute_changes_today_por_asignado(cache):
    """Entradas/salidas hoy agrupadas por (estado, persona) — usa person_for_status."""
    today = datetime.now(timezone.utc).date()
    cutoff_ayer = datetime(today.year, today.month, today.day, 0, 0, 0, tzinfo=timezone.utc) - timedelta(seconds=1)
    changes = {}
    for key, t in cache.get("tickets", {}).items():
        status_ayer = status_at(t, cutoff_ayer)
        status_hoy = t.get("current_status")
        if status_ayer == status_hoy:
            continue
        if status_ayer and is_open(status_ayer):
            p_ayer = person_for_status(t, status_ayer)
            k_out = (status_ayer, p_ayer)
            changes.setdefault(k_out, {"in": 0, "out": 0})["out"] += 1
        if status_hoy and is_open(status_hoy):
            p_hoy = person_for_status(t, status_hoy)
            k_in = (status_hoy, p_hoy)
            changes.setdefault(k_in, {"in": 0, "out": 0})["in"] += 1
    return changes


def build_tecnicos_section(cache, avances_extra=None):
    """Tabla estados x persona (assignee hasta Esperando inicio reparación, después tec_taller)."""
    grid = {}
    asignados_set = set()
    states_set = set()
    for key, t in cache.get("tickets", {}).items():
        if not is_open(t.get("current_status")):
            continue
        st = t.get("current_status") or ""
        persona = person_for_status(t, st)
        asignados_set.add(persona)
        states_set.add(st)
        grid[(st, persona)] = grid.get((st, persona), 0) + 1

    changes = compute_changes_today_por_asignado(cache)
    asignados_sorted = sorted(asignados_set, key=lambda x: (0 if x == "(Sin asignar)" else 1, x.lower()))
    states_order = [s for s in FUNNEL_ORDER if s in states_set] + sorted(s for s in states_set if s not in FUNNEL_ORDER)
    taller_set = set(TALLER_STATUSES_ORDER)

    ARROW_IN = "#c0392b"
    ARROW_OUT = "#1f8a4c"

    def flow_html(f_in, f_out):
        parts = []
        if f_in > 0:
            parts.append(f'<span style="color:{ARROW_IN};font-size:10px;font-weight:500" title="entran hoy">&#9650;{f_in}</span>')
        if f_out > 0:
            parts.append(f'<span style="color:{ARROW_OUT};font-size:10px;font-weight:500" title="salen hoy">&#9660;{f_out}</span>')
        return (' ' + ' '.join(parts)) if parts else ''

    # Totales por persona
    col_tot = {a: 0 for a in asignados_sorted}
    col_flow = {a: {"in": 0, "out": 0} for a in asignados_sorted}
    for (st, a), n in grid.items():
        if a in col_tot:
            col_tot[a] += n
    for (st, a), f in changes.items():
        if a in col_flow:
            col_flow[a]["in"] += f["in"]
            col_flow[a]["out"] += f["out"]
    grand_total = sum(col_tot.values())
    grand_flow = {"in": sum(f["in"] for f in col_flow.values()),
                  "out": sum(f["out"] for f in col_flow.values())}

    # Header (sin total grande — el total ya aparece en la fila Total inferior)
    parts = ['<tr><th class="sticky-l1 sticky-l2" style="text-align:left;left:0;min-width:240px">Estado</th>']
    for a in asignados_sorted:
        parts.append(
            f'<th style="padding:8px 10px"><div style="font-size:11px;line-height:1.2">{html_escape(a)}</div></th>'
        )
    parts.append('<th class="hdr-dia col-total" style="padding:8px 10px">Total</th>')
    parts.append('</tr>')
    header = "".join(parts)

    # Body
    body_parts = []
    for st in states_order:
        is_taller = st in taller_set
        idx = states_order.index(st)
        prev_taller = (idx > 0 and states_order[idx - 1] in taller_set)
        next_taller = (idx < len(states_order) - 1 and states_order[idx + 1] in taller_set)
        cls_list = []
        if is_taller:
            cls_list.append("row-taller")
            if not prev_taller:
                cls_list.append("row-taller-first")
            if not next_taller:
                cls_list.append("row-taller-last")
        if st not in ESPERANDO_TECNICO_STATES:
            cls_list.append("tec-no-aplica")
        cls_attr = f' class="{" ".join(cls_list)}"' if cls_list else ''
        row_tot = 0
        row_flow = {"in": 0, "out": 0}
        cells = [f'<tr{cls_attr}><td class="lbl-name" style="left:0;min-width:240px">{html_escape(st)}</td>']
        for a in asignados_sorted:
            n = grid.get((st, a), 0)
            f = changes.get((st, a), {"in": 0, "out": 0})
            row_tot += n
            row_flow["in"] += f["in"]
            row_flow["out"] += f["out"]
            if n > 0 or f["in"] > 0 or f["out"] > 0:
                content = f'<span style="font-weight:500">{n if n else "·"}</span>{flow_html(f["in"], f["out"])}'
                attrs = f' data-estado="{html_escape(st)}" data-asignado="{html_escape(a)}"'
                cells.append(f'<td class="num tec-cell"{attrs}>{content}</td>')
            else:
                cells.append('<td class="num n0"><span style="color:#c0d0e0">·</span></td>')
        # Total fila
        if row_tot > 0 or row_flow["in"] > 0 or row_flow["out"] > 0:
            content = f'<span style="font-weight:500">{row_tot}</span>{flow_html(row_flow["in"], row_flow["out"])}'
            attrs = f' data-estado="{html_escape(st)}" data-asignado=""'
            cells.append(f'<td class="num col-total tec-cell"{attrs}>{content}</td>')
        else:
            cells.append('<td class="num col-total n0"><span style="color:#c0d0e0">·</span></td>')
        cells.append('</tr>')
        body_parts.append("".join(cells))

    # Fila "Esperando técnico": suma de los 5 estados clave por persona
    esp_tot = {a: 0 for a in asignados_sorted}
    esp_flow = {a: {"in": 0, "out": 0} for a in asignados_sorted}
    for (st, a), n in grid.items():
        if st in ESPERANDO_TECNICO_STATES and a in esp_tot:
            esp_tot[a] += n
    for (st, a), f in changes.items():
        if st in ESPERANDO_TECNICO_STATES and a in esp_flow:
            esp_flow[a]["in"] += f["in"]
            esp_flow[a]["out"] += f["out"]
    esp_grand = sum(esp_tot.values())
    esp_grand_flow = {"in": sum(f["in"] for f in esp_flow.values()),
                      "out": sum(f["out"] for f in esp_flow.values())}
    esp_estado_list = "|".join(sorted(ESPERANDO_TECNICO_STATES))
    esp_cells = [f'<tr style="background:#fff4d6"><td class="lbl-name" style="left:0;min-width:240px;font-weight:500;background:#fff4d6" title="Suma de tickets en: {html_escape(esp_estado_list.replace("|", ", "))}">Esperando técnico</td>']
    for a in asignados_sorted:
        v = esp_tot.get(a, 0)
        f = esp_flow.get(a, {"in": 0, "out": 0})
        if v > 0 or f["in"] > 0 or f["out"] > 0:
            content = f'<span style="font-weight:600">{v}</span>{flow_html(f["in"], f["out"])}'
            attrs = f' data-estado-list="{html_escape(esp_estado_list)}" data-asignado="{html_escape(a)}"'
            esp_cells.append(f'<td class="num tec-cell" style="background:#fff4d6"{attrs}>{content}</td>')
        else:
            esp_cells.append('<td class="num n0" style="background:#fff4d6"><span style="color:#c0d0e0">·</span></td>')
    content = f'<span style="font-weight:600">{esp_grand}</span>{flow_html(esp_grand_flow["in"], esp_grand_flow["out"])}'
    esp_cells.append(f'<td class="num col-total tec-cell" style="background:#fff4d6;font-weight:600" data-estado-list="{html_escape(esp_estado_list)}" data-asignado="">{content}</td>')
    esp_cells.append('</tr>')
    body_parts.append("".join(esp_cells))

    # Fila Total
    tot_cells = ['<tr class="total"><td class="lbl-name" style="left:0;min-width:240px">Total</td>']
    for a in asignados_sorted:
        v = col_tot.get(a, 0)
        f = col_flow.get(a, {"in": 0, "out": 0})
        if v > 0 or f["in"] > 0 or f["out"] > 0:
            content = f'<span style="font-weight:500">{v}</span>{flow_html(f["in"], f["out"])}'
            attrs = f' data-estado="" data-asignado="{html_escape(a)}"'
            tot_cells.append(f'<td class="num tec-cell"{attrs}>{content}</td>')
        else:
            tot_cells.append('<td class="num n0"><span style="color:#c0d0e0">·</span></td>')
    content = f'<span style="font-weight:500">{grand_total}</span>{flow_html(grand_flow["in"], grand_flow["out"])}'
    tot_cells.append(f'<td class="num col-total tec-cell" data-estado="" data-asignado="">{content}</td>')
    tot_cells.append('</tr>')
    body_parts.append("".join(tot_cells))

    # === Filas adicionales de Avances (Presupuesto hecho, Equipo reparado) ===
    if avances_extra:
        # avances_extra = {"grid": {(label, persona): count}, "details": {(label, persona): [tickets]}, "labels": [(state, label)]}
        avg = avances_extra.get("grid", {})
        adetails = avances_extra.get("details", {})
        alabels = avances_extra.get("labels", [])
        for state_origin, lbl in alabels:
            row_cells = [f'<tr class="avances-row"><td class="lbl-name" style="left:0;min-width:240px;font-weight:500" title="Tickets unicos avanzados hoy desde {html_escape(state_origin)} hacia delante">{html_escape(lbl)} hoy</td>']
            row_tot = 0
            for a in asignados_sorted:
                n = avg.get((lbl, a), 0)
                row_tot += n
                if n > 0:
                    dkey = f"{lbl}__{a}".replace(" ", "_").replace("/", "_")
                    attrs = f' data-detail-key="{html_escape(dkey)}" data-estado="{html_escape(lbl)}" data-persona="{html_escape(a)}"'
                    row_cells.append(f'<td class="num avance-cell"{attrs}><span style="font-weight:500;color:#16a34a">{n}</span></td>')
                else:
                    row_cells.append('<td class="num n0"><span style="color:#c0d0e0">·</span></td>')
            # Total col
            if row_tot > 0:
                row_cells.append(f'<td class="num col-total"><span style="font-weight:600;color:#16a34a">{row_tot}</span></td>')
            else:
                row_cells.append('<td class="num col-total n0"><span style="color:#c0d0e0">·</span></td>')
            row_cells.append('</tr>')
            body_parts.append("".join(row_cells))

    return header, "\n".join(body_parts)


def build_avances_section(cache):
    """Métrica de productividad real:
    - "Presupuesto hecho" = transiciones hoy con to_status = "Presupuesto preparado pendiente de enviar"
    - "Equipo reparado" = transiciones hoy con to_status = "Inspección de salida"
    Dedupe por (ticket, persona): si un ticket toca el to_status 2 veces, cuenta 1 para esa persona.
    """
    # to_status objetivo -> label visible
    AVANCES_LABELS = [
        ("Presupuesto preparado pendiente de enviar", "Presupuesto hecho"),
        ("Inspección de salida", "Equipo reparado"),
    ]
    TO_STATES = {to for to, _ in AVANCES_LABELS}
    AVANCES_LABEL_BY_STATE = dict(AVANCES_LABELS)
    AVANCES_STATES_ORDER = [to for to, _ in AVANCES_LABELS]
    AVANCES_SET = set(AVANCES_STATES_ORDER)

    today = datetime.now(timezone.utc).date()
    cutoff = datetime(today.year, today.month, today.day, 0, 0, 0, tzinfo=timezone.utc)
    try:
        from zoneinfo import ZoneInfo
        madrid = ZoneInfo("Europe/Madrid")
    except Exception:
        madrid = None

    def fmt_hora(dt):
        try:
            return dt.astimezone(madrid).strftime("%H:%M") if madrid else dt.strftime("%H:%M")
        except Exception:
            return ""

    grid = {}
    detail_map = {}  # (to_state, person, ticket_key) -> {key, cliente, from, to, ts, dt}
    personas_set = set()

    for k, t in cache.get("tickets", {}).items():
        for tr in t.get("transitions", []):
            if len(tr) < 3:
                continue
            ts_str, from_s, to_s = tr[0], tr[1], tr[2]
            if to_s not in TO_STATES:
                continue
            dt = parse_iso(ts_str)
            if not dt or dt < cutoff:
                continue
            person = person_for_status(t, to_s)
            personas_set.add(person)
            cnt_key = (to_s, person, k)
            if cnt_key not in detail_map:
                grid[(to_s, person)] = grid.get((to_s, person), 0) + 1
            # Mantener la última transición a este to_state
            existing = detail_map.get(cnt_key)
            if not existing or dt > existing["dt"]:
                detail_map[cnt_key] = {
                    "key": k, "cliente": t.get("cliente", "") or "",
                    "from": from_s, "to": to_s,
                    "ts": fmt_hora(dt), "dt": dt,
                }

    details = {}
    for (to_s, person, k), info in detail_map.items():
        details.setdefault((to_s, person), []).append({
            "key": info["key"], "cliente": info["cliente"],
            "from": info["from"], "to": info["to"], "ts": info["ts"],
        })
    for lst in details.values():
        lst.sort(key=lambda x: x["ts"], reverse=True)

    if not personas_set:
        header = ('<tr><th class="sticky-l1 sticky-l2" style="text-align:left;left:0;min-width:240px">Tarea</th>'
                  '<th>(sin avances aún)</th><th class="col-total">Total</th></tr>')
        rows = ('<tr><td class="lbl-name" style="left:0;min-width:240px" colspan="3">'
                '<span style="color:#94a3b8;font-style:italic">Aún no hay tareas completadas hoy</span></td></tr>')
        return header, rows, "{}", {"grid": {}, "details": {}, "labels": []}

    personas_sorted = sorted(personas_set, key=lambda x: (0 if x == "(Sin asignar)" else 1, x.lower()))
    parts = ['<tr><th class="sticky-l1 sticky-l2" style="text-align:left;left:0;min-width:240px">Tarea</th>']
    for p in personas_sorted:
        parts.append(f'<th>{html_escape(p)}</th>')
    parts.append('<th class="hdr-dia col-total">Total</th></tr>')
    header = "".join(parts)
    body_parts = []
    col_tot = {p: 0 for p in personas_sorted}
    grand = 0
    detail_dict = {}
    for to_s in AVANCES_STATES_ORDER:
        lbl = AVANCES_LABEL_BY_STATE.get(to_s, to_s)
        row_tot = 0
        cells = [f'<tr><td class="lbl-name" style="left:0;min-width:240px">{html_escape(lbl)}</td>']
        for p in personas_sorted:
            n = grid.get((to_s, p), 0)
            row_tot += n
            col_tot[p] += n
            if n > 0:
                dkey = f"{lbl}__{p}".replace(" ", "_").replace("/", "_")
                detail_dict[dkey] = details.get((to_s, p), [])
                attrs = f' data-detail-key="{html_escape(dkey)}" data-estado="{html_escape(lbl)}" data-persona="{html_escape(p)}"'
                cells.append(f'<td class="num avance-cell"{attrs}><span style="font-weight:500">{n}</span></td>')
            else:
                cells.append('<td class="num n0"><span style="color:#c0d0e0">·</span></td>')
        grand += row_tot
        if row_tot > 0:
            cells.append(f'<td class="num col-total"><span style="font-weight:500">{row_tot}</span></td>')
        else:
            cells.append('<td class="num col-total n0"><span style="color:#c0d0e0">·</span></td>')
        cells.append('</tr>')
        body_parts.append("".join(cells))
    tot_cells = ['<tr class="total"><td class="lbl-name" style="left:0;min-width:240px">Total</td>']
    for p in personas_sorted:
        v = col_tot.get(p, 0)
        if v > 0:
            tot_cells.append(f'<td class="num"><span style="font-weight:600">{v}</span></td>')
        else:
            tot_cells.append('<td class="num n0"><span style="color:#c0d0e0">·</span></td>')
    tot_cells.append(f'<td class="num col-total"><span style="font-weight:600">{grand}</span></td>')
    tot_cells.append('</tr>')
    body_parts.append("".join(tot_cells))

    extra = {
        "grid": {(AVANCES_LABEL_BY_STATE.get(to, to), p): n for (to, p), n in grid.items()},
        "details": {(AVANCES_LABEL_BY_STATE.get(to, to), p): lst for (to, p), lst in details.items()},
        "labels": list(AVANCES_LABELS),
    }
    detail_dict_extra = {}
    for (to_s, p), lst in details.items():
        lbl = AVANCES_LABEL_BY_STATE.get(to_s, to_s)
        dkey = f"{lbl}__{p}".replace(" ", "_").replace("/", "_")
        detail_dict_extra[dkey] = lst
    detail_dict.update(detail_dict_extra)
    detail_json = json.dumps(detail_dict, ensure_ascii=False)
    return header, "\n".join(body_parts), detail_json, extra



_TT_JS = """<script>(function(){
  var tbl=document.getElementById('tt-table');if(!tbl)return;
  var rows=Array.from(tbl.tBodies[0].rows);
  var cnt=document.getElementById('tt-count');var tot=document.getElementById('tt-total');tot.textContent=rows.length;
  var fSearch=document.getElementById('tt-search');var fChain=document.getElementById('tt-chain');
  var fStatus=document.getElementById('tt-status');var fTipo=document.getElementById('tt-tipo');
  var fBloq=document.getElementById('tt-bloq');var fSusti=document.getElementById('tt-susti');
  var fBucket=document.getElementById('tt-bucket');
  function apply(){
    var q=(fSearch.value||'').toLowerCase().trim();
    var ch=fChain.value,st=fStatus.value,ti=fTipo.value,bl=fBloq.value,su=fSusti.value,bk=fBucket.value;
    var n=0;
    for(var i=0;i<rows.length;i++){var r=rows[i];
      var txt=r.textContent.toLowerCase();
      var days=parseInt(r.cells[10].textContent.trim())||0;
      var mb=true;
      if(bk==='g')mb=days<4;
      else if(bk==='g2')mb=days>=4&&days<=8;
      else if(bk==='y')mb=days>=9&&days<=15;
      else if(bk==='o')mb=days>=16&&days<=30;
      else if(bk==='r')mb=days>30;
      var ok=(!q||txt.indexOf(q)>=0)
        &&(!ch||r.cells[0].textContent.trim()===ch)
        &&(!st||r.cells[3].textContent.trim()===st)
        &&(!ti||r.cells[4].textContent.trim()===ti)
        &&(!bl||r.cells[7].textContent.trim()===bl)
        &&(!su||r.cells[8].textContent.trim()===su)
        &&mb;
      r.style.display=ok?'':'none';if(ok)n++;
    }cnt.textContent=n;
  }
  [fSearch,fChain,fStatus,fTipo,fBloq,fSusti,fBucket].forEach(function(e){e.addEventListener('input',apply);e.addEventListener('change',apply);});
  document.getElementById('tt-clear').addEventListener('click',function(){fSearch.value='';fChain.value='';fStatus.value='';fTipo.value='';fBloq.value='';fSusti.value='';fBucket.value='';apply();});
  var heads=tbl.tHead.rows[0].cells;var lastSort={col:-1,dir:1};
  for(var hi=0;hi<heads.length;hi++){(function(th){th.style.cursor='pointer';th.addEventListener('click',function(){
    var col=parseInt(th.dataset.col);var type=th.dataset.type||'text';var dir=(lastSort.col===col)?-lastSort.dir:1;lastSort={col:col,dir:dir};
    var rs=Array.from(tbl.tBodies[0].rows);
    rs.sort(function(a,b){var av=a.cells[col].textContent.trim();var bv=b.cells[col].textContent.trim();
      if(type==='num'){av=parseFloat(av.replace(/\\./g,''))||0;bv=parseFloat(bv.replace(/\\./g,''))||0;return(av-bv)*dir;}
      return av.localeCompare(bv,'es')*dir;});
    for(var k=0;k<rs.length;k++)tbl.tBodies[0].appendChild(rs[k]);
  });})(heads[hi]);}
  apply();
  var kpiTotal=document.getElementById('kpi-tiempo-taller-total');
  if(kpiTotal){kpiTotal.addEventListener('click',function(){
    fSearch.value='';fChain.value='';fStatus.value='';fTipo.value='';fBloq.value='';fSusti.value='';fBucket.value='';apply();
    document.getElementById('tt-table').scrollIntoView({behavior:'smooth',block:'start'});
  });}
})();</script>"""


def build_taller_detail_section(cache, sustis_by_parent=None):
    """Tabla detalle filtrable de tickets 2026 cerrados con tiempo en taller medible."""
    sustis_by_parent = sustis_by_parent or {}
    TALLER = {"Pendiente asignar técnico", "En cola taller",
              "En preparación presupuesto", "Esperando inicio reparación",
              "En reparación", "Inspección de salida"}
    EXIT = {"Devuelto a cliente", "Resuelto", "Finalizada", "Cancelado",
            "Finalizado técnico externo"}
    JIRA_URL = "https://leaseir.atlassian.net/browse/"
    rows = []
    for k, t in cache.get("tickets", {}).items():
        if is_open(t.get("current_status")):
            continue
        if not (t.get("created") or "").startswith("2026"):
            continue
        entrada = None
        salida = None
        for tr in t.get("transitions", []):
            if len(tr) < 3:
                continue
            ts, fs, to = tr[0], tr[1], tr[2]
            dt = parse_iso(ts)
            if not dt:
                continue
            if entrada is None and to == "Pendiente asignar técnico":
                entrada = dt
            if entrada and (to in EXIT) and (fs in TALLER):
                salida = dt
                break
        if not (entrada and salida):
            continue
        dias = (salida - entrada).days
        if dias < 0:
            continue
        chain = chain_of(t.get("cliente", ""), t.get("loc", ""))
        created_dt = parse_iso(t.get("created"))
        created_s = created_dt.strftime("%d/%m/%Y") if created_dt else ""
        salida_s = salida.strftime("%d/%m/%Y")
        resumen_av = ", ".join(t.get("motivo") or [])
        cambios_av = ", ".join(t.get("cambios") or [])
        rows.append({
            "chain": chain, "key": k,
            "cliente": t.get("cliente", ""),
            "status_final": t.get("current_status", ""),
            "tipo": t.get("tipo", ""),
            "resumen_av": resumen_av,
            "cambios_av": cambios_av,
            "bloq": t.get("bloq", ""),
            "susti": t.get("susti", ""),
            "susti_status": sustis_by_parent.get(k, ""),
            "dias_taller": dias,
            "consola": t.get("consola", ""),
            "hp": t.get("hp", ""),
            "garantia": t.get("garantia", ""),
            "fventa": t.get("fventa") or "",
            "disparos": t.get("disparos", ""),
            "importe": t.get("importe", ""),
            "created": created_s,
            "salida": salida_s,
            "asignado": t.get("asignado", ""),
        })
    if not rows:
        return '<div style="color:#94a3b8;padding:14px;font-style:italic">Sin tickets cerrados 2026 con tiempo medible.</div>'
    rows.sort(key=lambda r: -r["dias_taller"])

    def _dc(d):
        if d < 4:
            return "#1f8a4c"
        if d <= 8:
            return "#65a30d"
        if d <= 15:
            return "#b8860b"
        if d <= 30:
            return "#cb6f0a"
        return "#c0392b"

    chains_set = sorted({r["chain"] for r in rows})
    statuses_set = sorted({r["status_final"] for r in rows})
    tipos_set = sorted({r["tipo"] for r in rows if r["tipo"]})
    chain_opts = ''.join(f'<option value="{html_escape(c)}">{html_escape(c)}</option>' for c in chains_set)
    status_opts = ''.join(f'<option value="{html_escape(s)}">{html_escape(s)}</option>' for s in statuses_set)
    tipo_opts = ''.join(f'<option value="{html_escape(tp)}">{html_escape(tp)}</option>' for tp in tipos_set)
    out = []
    out.append(f'<div style="margin:8px 0 4px;color:var(--grey);font-size:13px"><b>{len(rows)}</b> tickets cerrados 2026 con tiempo en taller medible</div>')
    out.append('<div class="toolbar">')
    out.append('<input type="text" id="tt-search" placeholder="Buscar texto en cualquier columna...">')
    out.append(f'<select id="tt-chain"><option value="">Cadena: todas</option>{chain_opts}</select>')
    out.append(f'<select id="tt-status"><option value="">Estado: todos</option>{status_opts}</select>')
    out.append(f'<select id="tt-tipo"><option value="">Tipo avería: todos</option>{tipo_opts}</select>')
    out.append('<select id="tt-bloq"><option value="">Bloq: todos</option><option>Sí</option><option>No</option></select>')
    out.append('<select id="tt-susti"><option value="">Susti: todas</option><option>Sí</option><option>No</option></select>')
    out.append('<select id="tt-bucket"><option value="">Días: todos</option><option value="g">&lt;4d</option><option value="g2">5-8d</option><option value="y">9-15d</option><option value="o">16-30d</option><option value="r">&gt;30d</option></select>')
    out.append('<button type="button" id="tt-clear" class="multi-btn">✕ Limpiar</button>')
    out.append('<span class="count"><b id="tt-count">0</b> de <span id="tt-total">0</span> visibles</span>')
    out.append('</div>')
    out.append('<div class="scroller"><table id="tt-table" style="font-size:11px;min-width:1700px"><thead><tr>')
    hdrs = [
        ("Cadena", "text"), ("Ticket", "text"), ("Cliente", "text"),
        ("Estado final", "text"), ("Tipo avería", "text"),
        ("Resumen avería", "text"), ("Cambios/Mejoras", "text"),
        ("Bloq", "text"), ("Susti", "text"), ("Estado susti", "text"),
        ("Días en taller", "num"),
        ("Consola", "text"), ("HP", "text"), ("Garantía", "text"),
        ("Disparos", "num"), ("Importe", "num"),
        ("Creada", "date"), ("Salida", "date"), ("Persona", "text"),
    ]
    for i, (h, dtype) in enumerate(hdrs):
        out.append(f'<th class="sortable" data-col="{i}" data-type="{dtype}">{html_escape(h)}</th>')
    out.append('</tr></thead><tbody>')
    for r in rows:
        link = f'<a href="{JIRA_URL}{r["key"]}" target="_blank" rel="noopener" style="color:#2a59c4;text-decoration:none">{r["key"]}</a>'
        dcolor = _dc(r["dias_taller"])
        out.append('<tr>')
        out.append(f'<td>{html_escape(r["chain"])}</td>')
        out.append(f'<td>{link}</td>')
        out.append(f'<td class="d-trunc" title="{html_escape(r["cliente"])}">{html_escape(r["cliente"])}</td>')
        out.append(f'<td>{html_escape(r["status_final"])}</td>')
        out.append(f'<td>{html_escape(r["tipo"])}</td>')
        out.append(f'<td class="d-trunc" title="{html_escape(r["resumen_av"])}">{html_escape(r["resumen_av"])}</td>')
        out.append(f'<td class="d-trunc" title="{html_escape(r["cambios_av"])}">{html_escape(r["cambios_av"])}</td>')
        out.append(f'<td style="text-align:center">{html_escape(r["bloq"])}</td>')
        out.append(f'<td style="text-align:center">{html_escape(r["susti"])}</td>')
        out.append(f'<td style="text-align:center;font-size:11px">{html_escape(r["susti_status"]) if r["susti_status"] else "—"}</td>')
        out.append(f'<td style="text-align:center;color:{dcolor};font-weight:600">{r["dias_taller"]}</td>')
        out.append(f'<td>{html_escape(r["consola"])}</td>')
        out.append(f'<td>{html_escape(r["hp"])}</td>')
        out.append(f'<td style="text-align:center">{html_escape(r["garantia"])}</td>')
        out.append(f'<td style="text-align:right;font-variant-numeric:tabular-nums">{html_escape(format_int_dot(r["disparos"]))}</td>')
        out.append(f'<td style="text-align:right">{html_escape(str(r["importe"]))}</td>')
        out.append(f'<td>{r["created"]}</td>')
        out.append(f'<td>{r["salida"]}</td>')
        out.append(f'<td>{html_escape(r["asignado"])}</td>')
        out.append('</tr>')
    out.append('</tbody></table></div>')
    out.append(_TT_JS)
    return "".join(out)




def build_tiempo_sustis_section(out_path):
    """Stacked bar de tiempo de sustituciones ACTIVAS (días desde fecha_envio).
    Aprovecha sustis_activas.json. Las sustis cerradas no están aquí, así que
    medimos el tiempo CURRENT de cada sustitución activa."""
    BUCKETS = [
        ("0-7 días",   "#1f8a4c", lambda d: d <= 7),
        ("8-15 días",  "#65a30d", lambda d: 8 <= d <= 15),
        ("16-25 días", "#b8860b", lambda d: 16 <= d <= 25),
        ("26-35 días", "#cb6f0a", lambda d: 26 <= d <= 35),
        (">35 días",   "#c0392b", lambda d: d > 35),
    ]
    sustis_path = None
    try:
        from pathlib import Path as _P
        for _cand in [_P("cache") / "sustis_activas.json", out_path.parent.parent / "cache" / "sustis_activas.json"]:
            if _cand.exists():
                sustis_path = _cand
                break
    except Exception:
        pass
    if not sustis_path:
        return '<div style="color:#94a3b8;padding:14px;font-style:italic">Sin datos de sustituciones disponibles.</div>'
    try:
        data = json.loads(sustis_path.read_text(encoding="utf-8"))
    except Exception as e:
        return f'<div style="color:#c0392b;padding:14px">Error leyendo sustis: {e}</div>'

    now = datetime.now(timezone.utc)
    counts = [0] * len(BUCKETS)
    dias_list = []
    sample_max = []
    for it in (data.get("items") or []):
        fe = it.get("fecha_envio")
        if not fe: continue
        fe_dt = parse_iso(fe)
        if not fe_dt: continue
        if not fe_dt.tzinfo:
            fe_dt = fe_dt.replace(tzinfo=timezone.utc)
        dias = (now - fe_dt).days
        if dias < 0: continue
        dias_list.append(dias)
        sample_max.append((dias, it.get("parent_key") or it.get("key") or ""))
        for i, (_, _, pred) in enumerate(BUCKETS):
            if pred(dias):
                counts[i] += 1
                break
    total = sum(counts)
    if not total:
        return '<div style="color:#94a3b8;padding:14px;font-style:italic">No hay sustituciones activas con fecha de envío.</div>'
    media = sum(dias_list) / total
    mediana = sorted(dias_list)[total // 2]
    max_d = max(dias_list)
    sample_max.sort(reverse=True)
    top3 = sample_max[:3]
    segments = []
    for (lbl, color, _), n in zip(BUCKETS, counts):
        pct = (n * 100.0 / total) if total else 0
        if n > 0:
            label_inline = "" if pct < 5 else str(n)
            segments.append(
                f'<div style="background:{color};height:100%;width:{pct:.2f}%;display:flex;align-items:center;justify-content:center;color:white;font-weight:600;font-size:13px;min-width:0;overflow:hidden" title="{lbl}: {n} sustis ({pct:.0f}%)">{label_inline}</div>'
            )
    bar_html = '<div style="display:flex;height:36px;border-radius:6px;overflow:hidden;border:1px solid var(--line);background:#f1f5f9">' + "".join(segments) + '</div>'
    legend_items = []
    for (lbl, color, _), n in zip(BUCKETS, counts):
        pct = (n * 100.0 / total) if total else 0
        legend_items.append(
            f'<span style="display:inline-flex;align-items:center;gap:6px;margin-right:14px;font-size:13px">'
            f'<span style="display:inline-block;width:12px;height:12px;background:{color};border-radius:3px"></span>'
            f'<b>{lbl}</b>: {n} ({pct:.0f}%)</span>'
        )
    top_label = html_escape(top3[0][1]) if top3 else ""
    kpis = (
        '<div style="display:flex;gap:12px;margin:6px 0 12px;flex-wrap:wrap">'
        f'<div class="kpi" style="flex:1 1 150px"><div class="k">Total sustis activas</div><div class="v">{total}</div><div class="d">con fecha de envío</div></div>'
        f'<div class="kpi" style="flex:1 1 150px"><div class="k">Mediana</div><div class="v">{mediana:.0f}<span style="font-size:14px;color:var(--grey);margin-left:4px">días</span></div><div class="d">tiempo actual de préstamo</div></div>'
        f'<div class="kpi" style="flex:1 1 150px"><div class="k">Media</div><div class="v">{media:.1f}<span style="font-size:14px;color:var(--grey);margin-left:4px">días</span></div><div class="d">sensible a outliers</div></div>'
        f'<div class="kpi" style="flex:1 1 150px"><div class="k">Máximo</div><div class="v" style="color:#c0392b">{max_d}<span style="font-size:14px;color:var(--grey);margin-left:4px">días</span></div><div class="d">{top_label}</div></div>'
        '</div>'
    )
    return (
        kpis + bar_html +
        '<div style="margin-top:12px">' + "".join(legend_items) + '</div>'
        '<div class="legend" style="margin-top:8px">Tiempo desde <b>fecha y hora de envío</b> de la sub-tarea de sustitución hasta hoy. Solo sustis activas (cuando se devuelven dejan de aparecer en el cache).</div>'
    )


def build_anual_averias_section(cache):
    """Tabla cruzada Año venta x Tipo averia para tickets 2026 que pasaron por
    Inspeccion de salida. Usa customfield_10815 (Cambios obs y mejoras),
    agrupado en 10 buckets clinicos. Genera datos embebidos en JS para que el
    filtro (cadena + importe) se aplique client-side sin recargar la pagina.
    """
    INSP_STATES = {"Inspeccion de salida", "Inspección de salida"}

    BUCKETS = [
        ("Umbi", "Umbilical",
         ["Cambio de umbilical", "Reparación umbilical", "Reparacion umbilical"]),
        ("Punt", "Puntera/Óptica",
         ["Cambio de puntera", "Cambio de zafiro", "Cambio de lente trasera", "Cambio de prisma"]),
        ("Buff", "Buffer", ["Cambio de buffer"]),
        ("Diod", "Diodo", ["Diodo nuevo", "Diodo antirretorno"]),
        ("Plac", "Placa/Electr.",
         ["Cambio de placa de carga", "MOSFET de disparo 1º", "MOSFET de disparo 2º",
          "MOSFET de carga", "Integrado LT1054", "Cambio de driver", "Pila BIOS",
          "Unidad SSD", "RAM"]),
        ("Gati", "Gatillo/Pistola",
         ["Cambio switch de gatillo", "Cambio de carcasa de pistola"]),
        ("Pant", "Pantalla/Control",
         ["Cambio de táctil + controlador", "Cambio de tactil + controlador",
          "Cambio de control", "Switch de pedal"]),
        ("Refr", "Refrigeración",
         ["Cambio de nevera", "Cambio de termostato", "Ventilador de 12V",
          "Mejora de hidráulica", "Mejora de hidraulica"]),
        ("Mant", "Mantenimiento",
         ["Limpieza interior y exterior", "Llenado de depósito", "Llenado de deposito",
          "Calibrado y medición de energías", "Calibrado y medicion de energias",
          "Mejora de masas", "Actualizar software", "Casquillo nuevo", "Lanyard", "Buzzer"]),
        ("Otro", "Otros", ["Otros"]),
    ]
    BUCKET_KEYS = [b[0] for b in BUCKETS]
    BUCKET_LABELS = {b[0]: b[1] for b in BUCKETS}
    val2bucket = {}
    for short, lbl, vals in BUCKETS:
        for v in vals:
            val2bucket[v.lower()] = short

    def importe_bucket(imp):
        try:
            v = float(imp) if imp not in ("", None) else 0
        except (TypeError, ValueError):
            v = 0
        if v == 0: return "free"
        if v < 1000: return "low"
        if v <= 6000: return "mid"
        return "high"

    YEAR_ROWS = ["<=2018", "2019", "2020", "2021", "2022", "2023", "2024", "2025+", "Desconocido",
                 "Abierta", "Externa (técnico ext.)", "Cerrado s/insp", "Cancelado"]
    def year_row(y):
        if not y: return "Desconocido"
        try:
            yi = int(y)
        except (TypeError, ValueError):
            return "Desconocido"
        if yi <= 1999: return "Desconocido"
        if yi <= 2018: return "<=2018"
        if yi >= 2025: return "2025+"
        return str(yi)

    # Cargar serial_year para unificar con la lógica de "Por cadena"
    serial_year = {}
    try:
        from pathlib import Path as _PSY
        for _cand in [_PSY("data") / "serial_year.json", Path("data") / "serial_year.json"]:
            if _cand.exists():
                serial_year = json.loads(_cand.read_text(encoding="utf-8"))
                break
    except Exception:
        pass
    def _normalize_serial(s):
        if not s: return ""
        import re as _re
        s = _re.sub(r"\s+", "", str(s)).upper()
        for prefix in ("C", "H", "MHR", "AHR", "MHP", "HP"):
            if s.startswith(prefix) and len(s) > len(prefix) and s[len(prefix):].isdigit():
                return s[len(prefix):].lstrip("0") or "0"
        if s.isdigit(): return s.lstrip("0") or "0"
        return s
    def _year_from_serial(t):
        for f in ("consola", "hp"):
            s = (t.get(f) or "").strip()
            if not s: continue
            ns = _normalize_serial(s)
            for k_try in (s.upper(), ns):
                if k_try in serial_year:
                    return serial_year[k_try]
        return None

    records = []
    for k, t in cache.get("tickets", {}).items():
        created = t.get("created") or ""
        if not created.startswith("2026"):
            continue
        passed = False
        for tr in t.get("transitions", []):
            if len(tr) >= 3 and tr[2] in INSP_STATES:
                passed = True
                break
        is_currently_open = is_open(t.get("current_status"))
        st_now = t.get("current_status", "")
        # Año compra: serial_year primero (más fiable), fallback a fventa
        yr = _year_from_serial(t)
        if not yr:
            fv = (t.get("fventa") or "").strip()
            if fv and len(fv) >= 4 and fv[:4].isdigit():
                yr = int(fv[:4])
        buckets = set()
        if passed:
            for v in (t.get("cambios") or []):
                b = val2bucket.get(str(v).lower())
                if b: buckets.add(b)
            # Complementar SIEMPRE con resumen avería para Diodo/Umbilical
            # (a veces no se rellenan bien en cambios pero sí en motivo)
            for v in (t.get("motivo") or []):
                vl = str(v).lower()
                if "umbilical" in vl: buckets.add("Umbi")
                elif "diodo" in vl: buckets.add("Diod")
            # Fallback completo si todavía no hay buckets
            if not buckets:
                for v in (t.get("motivo") or []):
                    vl = str(v).lower()
                    if "umbilical" in vl: buckets.add("Umbi")
                    elif "buffer" in vl: buckets.add("Buff")
                    elif "puntera" in vl or "zafiro" in vl: buckets.add("Punt")
                    elif "diodo" in vl: buckets.add("Diod")
                    elif "placa" in vl: buckets.add("Plac")
                    elif "gatillo" in vl: buckets.add("Gati")
                    elif "pantalla" in vl: buckets.add("Pant")
                    elif "otros" in vl: buckets.add("Otro")
        ch = chain_of(t.get("cliente", ""), t.get("loc", ""))
        # PRIORIDAD: si sigue abierto hoy, es "Abierta" (aunque haya pasado por Inspección).
        if is_currently_open:
            yrow = "Abierta"
        elif passed:
            yrow = year_row(yr)
        elif st_now == "Cancelado":
            yrow = "Cancelado"
        elif st_now == "Finalizado técnico externo":
            yrow = "Externa (técnico ext.)"
        else:
            yrow = "Cerrado s/insp"
        records.append({"k": k, "y": yrow, "c": ch,
                        "i": importe_bucket(t.get("importe")),
                        "b": sorted(buckets)})

    if not records:
        return ('<div style="color:#94a3b8;padding:14px;font-style:italic">'
                'Aun no hay reparaciones 2026 con datos suficientes.</div>')

    chains_present = set(r["c"] for r in records)
    cadenas = [c for c in CHAIN_ORDER if c in chains_present]
    n_total = len(records)

    data_json = json.dumps(records, ensure_ascii=False)
    cad_options = ''.join(f'<option value="{html_escape(c)}">{html_escape(c)}</option>' for c in cadenas)
    bucket_headers = ''.join(
        f'<th style="background:var(--blue);color:white;font-weight:500;padding:6px 8px;text-align:right" title="{html_escape(BUCKET_LABELS[k])}">{k}</th>'
        for k in BUCKET_KEYS)
    year_rows_html = ''.join(
        f'<tr data-year="{y}"><td style="padding:6px 10px;font-weight:500">{y}</td>'
        f'<td class="num tickets-cell">0</td>'
        + ''.join(f'<td class="num bucket-cell" data-bucket="{k}">0</td>' for k in BUCKET_KEYS)
        + '</tr>'
        for y in YEAR_ROWS)

    html = (
        '<div style="display:flex;gap:14px;align-items:center;flex-wrap:wrap;margin:6px 0 10px;padding:9px 12px;background:#f1f5f9;border:1px solid var(--line);border-radius:7px;font-size:12px">'
        '<label style="display:inline-flex;align-items:center;gap:6px"><span style="color:#475569">Cadena:</span>'
        '<select id="av-filter-chain" style="padding:4px 8px;border:1px solid var(--line);border-radius:4px;font-size:12px;background:white">'
        '<option value="">Todas</option>'
        f'{cad_options}'
        '</select></label>'
        '<label style="display:inline-flex;align-items:center;gap:6px"><span style="color:#475569">Importe:</span>'
        '<select id="av-filter-importe" style="padding:4px 8px;border:1px solid var(--line);border-radius:4px;font-size:12px;background:white">'
        '<option value="">Todos</option>'
        '<option value="free">Sin coste</option>'
        '<option value="low">&lt; 1.000 €</option>'
        '<option value="mid">1.000 - 6.000 €</option>'
        '<option value="high">&gt; 6.000 €</option>'
        '</select></label>'
        '<span style="margin-left:auto;color:#475569">Tickets visibles: <b id="av-count" style="color:var(--blue)">'
        f'{n_total}</b> / {n_total}</span>'
        '</div>'
        '<div style="border:1px solid var(--line);border-radius:8px;overflow:hidden;background:white">'
        '<table id="av-table" style="width:100%;border-collapse:collapse;font-size:12.5px">'
        '<thead><tr>'
        '<th style="background:var(--blue);color:white;font-weight:500;padding:6px 10px;text-align:left">Año venta</th>'
        '<th style="background:var(--blue);color:white;font-weight:500;padding:6px 8px;text-align:right">Tickets</th>'
        f'{bucket_headers}'
        '</tr></thead>'
        f'<tbody>{year_rows_html}</tbody>'
        '<tfoot><tr class="total"><td style="padding:6px 10px;font-weight:600">TOTAL</td>'
        '<td class="num tickets-cell" style="font-weight:600">0</td>'
        + ''.join(f'<td class="num bucket-cell" data-bucket="{k}" style="font-weight:600">0</td>' for k in BUCKET_KEYS)
        + '</tr></tfoot>'
        '</table></div>'
        '<div class="legend" style="margin-top:6px">'
        + ' &middot; '.join(f'<b>{k}</b>={html_escape(BUCKET_LABELS[k])}' for k in BUCKET_KEYS) +
        '</div>'
        '<script id="av-data" type="application/json">' + data_json + '</script>'
        '<script>'
        '(function(){'
        'var DATA = JSON.parse(document.getElementById("av-data").textContent);'
        'var YEARS = ' + json.dumps(YEAR_ROWS) + ';'
        'var BUCKETS = ' + json.dumps(BUCKET_KEYS) + ';'
        'var tbl = document.getElementById("av-table");'
        'var selC = document.getElementById("av-filter-chain");'
        'var selI = document.getElementById("av-filter-importe");'
        'var cnt = document.getElementById("av-count");'
        'function recompute(){'
        '  var ch = selC.value, im = selI.value;'
        '  var grid = {}; YEARS.forEach(function(y){grid[y]={"_t":0}; BUCKETS.forEach(function(b){grid[y][b]=0;});});'
        '  var totals = {"_t":0}; BUCKETS.forEach(function(b){totals[b]=0;});'
        '  var visible = 0;'
        '  DATA.forEach(function(r){'
        '    if(ch && r.c!==ch) return;'
        '    if(im && r.i!==im) return;'
        '    visible++; grid[r.y]._t++; totals._t++;'
        '    r.b.forEach(function(b){ if(grid[r.y][b]!==undefined){grid[r.y][b]++; totals[b]++;} });'
        '  });'
        '  cnt.textContent = visible;'
        '  tbl.querySelectorAll("tbody tr").forEach(function(tr){'
        '    var y = tr.getAttribute("data-year"); var t = grid[y]._t;'
        '    var tc = tr.querySelector(".tickets-cell");'
        '    tc.textContent = t || "·"; tc.style.color = t ? "" : "#c0d0e0";'
        '    tr.querySelectorAll(".bucket-cell").forEach(function(td){'
        '      var v = grid[y][td.getAttribute("data-bucket")];'
        '      td.textContent = v || "·"; td.style.color = v ? "" : "#c0d0e0";'
        '    });'
        '  });'
        '  var tf = tbl.querySelector("tfoot tr");'
        '  tf.querySelector(".tickets-cell").textContent = totals._t;'
        '  tf.querySelectorAll(".bucket-cell").forEach(function(td){'
        '    td.textContent = totals[td.getAttribute("data-bucket")] || "·";'
        '  });'
        '}'
        'selC.addEventListener("change", recompute);'
        'selI.addEventListener("change", recompute);'
        'recompute();'
        '})();'
        '</script>'
    )
    return html




# ========== TIEMPOS REFACTORIZADOS ==========
# Base = creados 2026 OR (creados antes Y abiertos al 1-ene-2026)

CUTOFF_2026 = datetime(2026, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
TALLER_TRANSIT_STATES = {"Pendiente asignar técnico", "En cola taller",
                          "En preparación presupuesto", "Esperando inicio reparación",
                          "En reparación", "Inspección de salida"}
EXIT_TRANSIT_STATES = {"Devuelto a cliente", "Resuelto", "Finalizada", "Cancelado",
                        "Finalizado técnico externo"}


def is_in_universe_2026(ticket):
    """True si el ticket entra en la base unificada de Tiempos."""
    created = ticket.get("created") or ""
    if created.startswith("2026"):
        return True
    # Creado antes pero abierto al 1-ene-2026
    status_at_cutoff = status_at(ticket, CUTOFF_2026)
    if status_at_cutoff and status_at_cutoff in OPEN_STATUSES:
        return True
    return False


def compute_tiempo_taller(ticket, now_utc):
    """Devuelve dict con entrada, salida, dias, live (bool si aún en taller/abierto)."""
    entrada = None
    salida = None
    for tr in ticket.get("transitions", []):
        if len(tr) < 3:
            continue
        ts, fs, to = tr[0], tr[1], tr[2]
        dt = parse_iso(ts)
        if not dt:
            continue
        if entrada is None and to == "Pendiente asignar técnico":
            entrada = dt
        if entrada and (to in EXIT_TRANSIT_STATES) and (fs in TALLER_TRANSIT_STATES):
            salida = dt
            break
    if not entrada:
        return None
    live = False
    if salida:
        dias = (salida - entrada).days
    else:
        # Aún en taller o en otro estado abierto: tiempo desde entrada hasta ahora
        if is_open(ticket.get("current_status")):
            dias = (now_utc - entrada).days
            live = True
        else:
            # Cerrado pero sin salida limpia detectada → usar última transition
            last = ticket.get("transitions", [])
            if last:
                last_dt = parse_iso(last[-1][0])
                dias = (last_dt - entrada).days if last_dt else 0
            else:
                dias = 0
    if dias < 0:
        dias = 0
    return {"entrada": entrada, "salida": salida, "dias": dias, "live": live}


def collect_tiempos_data(cache, sustis_by_parent, sustis_envio_by_parent):
    """Recorre el universo y devuelve filas con toda la info para tabla detalle.
    sustis_envio_by_parent: dict parent_key -> fecha_envio_dt (para días susti)."""
    now_utc = datetime.now(timezone.utc)
    rows = []
    for k, t in cache.get("tickets", {}).items():
        if not is_in_universe_2026(t):
            continue
        taller = compute_tiempo_taller(t, now_utc)
        # Días susti
        susti_status = sustis_by_parent.get(k, "")
        fe = sustis_envio_by_parent.get(k)
        dias_susti = None
        if fe:
            if not fe.tzinfo:
                fe = fe.replace(tzinfo=timezone.utc)
            dias_susti = (now_utc - fe).days
        rows.append({
            "k": k,
            "chain": chain_of(t.get("cliente", ""), t.get("loc", "")),
            "cliente": t.get("cliente", ""),
            "status": t.get("current_status", ""),
            "is_open": is_open(t.get("current_status", "")),
            "tipo": t.get("tipo", ""),
            "motivo": ", ".join(t.get("motivo") or []),
            "cambios": ", ".join(t.get("cambios") or []),
            "bloq": t.get("bloq", ""),
            "susti": t.get("susti", ""),
            "susti_status": susti_status,
            "consola": t.get("consola", ""),
            "hp": t.get("hp", ""),
            "garantia": t.get("garantia", ""),
            "fventa": t.get("fventa") or "",
            "disparos": t.get("disparos", ""),
            "importe": t.get("importe", ""),
            "asignado": t.get("asignado", ""),
            "tec_taller": t.get("tec_taller", ""),
            "tec_externo": t.get("tec_externo", ""),
            "dias_taller": taller["dias"] if taller else None,
            "taller_live": taller["live"] if taller else False,
            "entrada_taller": taller["entrada"].strftime("%d/%m/%Y") if taller and taller["entrada"] else "",
            "salida_taller": taller["salida"].strftime("%d/%m/%Y") if taller and taller["salida"] else "",
            "fecha_envio_susti": fe.strftime("%d/%m/%Y") if fe else "",
            "dias_susti": dias_susti,
            "created": (parse_iso(t.get("created")) or now_utc).strftime("%d/%m/%Y"),
        })
    return rows


def _bar_html(buckets, counts, total, label, kpi_id="taller", ranges=None):
    """ranges: lista de (min, max) en días para cada bucket — se inyecta como data-tiempos-bucket."""
    segments = []
    for i, ((lbl, color, _), n) in enumerate(zip(buckets, counts)):
        pct = (n * 100.0 / total) if total else 0
        if n > 0:
            label_inline = "" if pct < 5 else str(n)
            bucket_attr = ""
            if ranges and i < len(ranges):
                lo, hi = ranges[i]
                bucket_attr = f' data-tiempos-bucket="{lo},{hi}" data-tiempos-kpi="{kpi_id}"'
            segments.append(
                f'<div{bucket_attr} style="background:{color};height:100%;width:{pct:.2f}%;display:flex;align-items:center;justify-content:center;color:white;font-weight:600;font-size:13px;min-width:0;overflow:hidden;cursor:pointer" title="{lbl}: {n} ({pct:.0f}%) — clic para filtrar">{label_inline}</div>'
            )
    return '<div style="display:flex;height:36px;border-radius:6px;overflow:hidden;border:1px solid var(--line);background:#f1f5f9">' + "".join(segments) + '</div>'


def _kpis_html(prefix, total, mediana, media, max_d, kpi_id_prefix, label_total):
    return (
        '<div style="display:flex;gap:12px;margin:6px 0 12px;flex-wrap:wrap">'
        f'<div class="kpi clickable" data-tiempos-filter="all" data-tiempos-kpi="{kpi_id_prefix}" style="flex:1 1 150px;cursor:pointer"><div class="k">Total tickets</div><div class="v">{total}</div><div class="d">{label_total}</div></div>'
        f'<div class="kpi clickable" data-tiempos-filter="mediana" data-tiempos-kpi="{kpi_id_prefix}" style="flex:1 1 150px;cursor:pointer"><div class="k">Mediana</div><div class="v">{mediana:.0f}<span style="font-size:14px;color:var(--grey);margin-left:4px">días</span></div><div class="d">Click: ordena por días</div></div>'
        f'<div class="kpi clickable" data-tiempos-filter="media" data-tiempos-kpi="{kpi_id_prefix}" style="flex:1 1 150px;cursor:pointer"><div class="k">Media</div><div class="v">{media:.1f}<span style="font-size:14px;color:var(--grey);margin-left:4px">días</span></div><div class="d">Click: ordena por días</div></div>'
        f'<div class="kpi clickable" data-tiempos-filter="max" data-tiempos-kpi="{kpi_id_prefix}" style="flex:1 1 150px;cursor:pointer"><div class="k">Máximo</div><div class="v" style="color:#c0392b">{max_d}<span style="font-size:14px;color:var(--grey);margin-left:4px">días</span></div><div class="d">Click: solo top 10%</div></div>'
        '</div>'
    )


def build_tiempos_section(cache, sustis_by_parent):
    """Devuelve dict con html para Tiempo taller, Tiempo sustis y Detalle compartido."""
    # Cargar fechas de envío de sustis activas
    sustis_envio_by_parent = {}
    try:
        from pathlib import Path as _PS
        for _cand in [_PS("cache") / "sustis_activas.json", Path("cache") / "sustis_activas.json"]:
            if _cand.exists():
                _su = json.loads(_cand.read_text(encoding="utf-8"))
                for _it in _su.get("items", []) or []:
                    _pk = _it.get("parent_key")
                    _fe = _it.get("fecha_envio")
                    if _pk and _fe:
                        _dt = parse_iso(_fe)
                        if _dt:
                            sustis_envio_by_parent[_pk] = _dt
                break
    except Exception:
        pass

    rows = collect_tiempos_data(cache, sustis_by_parent, sustis_envio_by_parent)
    if not rows:
        empty = '<div style="color:#94a3b8;padding:14px;font-style:italic">Sin tickets en la base.</div>'
        return {"taller": empty, "sustis": empty, "detail": empty}

    # === Tiempo taller ===
    taller_buckets = [
        ("<4 días",   "#1f8a4c", lambda d: d < 4),
        ("5-8 días",  "#65a30d", lambda d: 4 <= d <= 8),
        ("9-15 días", "#b8860b", lambda d: 9 <= d <= 15),
        ("16-30 días","#cb6f0a", lambda d: 16 <= d <= 30),
        (">30 días",  "#c0392b", lambda d: d > 30),
    ]
    taller_dias = [r["dias_taller"] for r in rows if r["dias_taller"] is not None]
    counts_taller = [0] * len(taller_buckets)
    for d in taller_dias:
        for i, (_, _, pred) in enumerate(taller_buckets):
            if pred(d):
                counts_taller[i] += 1
                break
    t_total = len(taller_dias)
    t_mediana = sorted(taller_dias)[t_total // 2] if t_total else 0
    t_media = (sum(taller_dias) / t_total) if t_total else 0
    t_max = max(taller_dias) if taller_dias else 0
    n_live = sum(1 for r in rows if r["taller_live"])
    taller_kpis = _kpis_html("t", t_total, t_mediana, t_media, t_max, "taller",
                             f"{n_live} aún en taller (live)")
    taller_ranges = [(0,3),(4,8),(9,15),(16,30),(31,9999)]
    taller_bar = _bar_html(taller_buckets, counts_taller, t_total, "taller", "taller", taller_ranges)
    taller_legend = ""
    for (lbl, color, _), n in zip(taller_buckets, counts_taller):
        pct = (n * 100.0 / t_total) if t_total else 0
        taller_legend += (f'<span style="display:inline-flex;align-items:center;gap:6px;margin-right:14px;font-size:13px">'
                          f'<span style="display:inline-block;width:12px;height:12px;background:{color};border-radius:3px"></span>'
                          f'<b>{lbl}</b>: {n} ({pct:.0f}%)</span>')
    taller_html = (taller_kpis + taller_bar +
                   '<div style="margin-top:12px">' + taller_legend + '</div>'
                   '<div class="legend" style="margin-top:8px">Base: tickets 2026 + abiertos al 1-ene-2026. Para abiertos aún en taller, tiempo = entrada a Pdte. asignar técnico hasta HOY.</div>')

    # === Tiempo sustituciones ===
    susti_buckets = [
        ("0-7 días",   "#1f8a4c", lambda d: d <= 7),
        ("8-15 días",  "#65a30d", lambda d: 8 <= d <= 15),
        ("16-25 días", "#b8860b", lambda d: 16 <= d <= 25),
        ("26-35 días", "#cb6f0a", lambda d: 26 <= d <= 35),
        (">35 días",   "#c0392b", lambda d: d > 35),
    ]
    susti_dias = [r["dias_susti"] for r in rows if r["dias_susti"] is not None]
    counts_susti = [0] * len(susti_buckets)
    for d in susti_dias:
        for i, (_, _, pred) in enumerate(susti_buckets):
            if pred(d):
                counts_susti[i] += 1
                break
    s_total = len(susti_dias)
    if s_total:
        s_mediana = sorted(susti_dias)[s_total // 2]
        s_media = sum(susti_dias) / s_total
        s_max = max(susti_dias)
    else:
        s_mediana = s_media = s_max = 0
    sustis_kpis = _kpis_html("s", s_total, s_mediana, s_media, s_max, "sustis",
                             "sustituciones con fecha de envío conocida")
    sustis_ranges = [(0,7),(8,15),(16,25),(26,35),(36,9999)]
    sustis_bar = _bar_html(susti_buckets, counts_susti, s_total, "sustis", "sustis", sustis_ranges)
    sustis_legend = ""
    for (lbl, color, _), n in zip(susti_buckets, counts_susti):
        pct = (n * 100.0 / s_total) if s_total else 0
        sustis_legend += (f'<span style="display:inline-flex;align-items:center;gap:6px;margin-right:14px;font-size:13px">'
                          f'<span style="display:inline-block;width:12px;height:12px;background:{color};border-radius:3px"></span>'
                          f'<b>{lbl}</b>: {n} ({pct:.0f}%)</span>')
    if s_total == 0:
        sustis_html = '<div style="color:#94a3b8;padding:14px;font-style:italic">No hay sustituciones activas con fecha de envío en la base 2026.</div>'
    else:
        sustis_html = (sustis_kpis + sustis_bar +
                       '<div style="margin-top:12px">' + sustis_legend + '</div>'
                       '<div class="legend" style="margin-top:8px">Tiempo desde fecha de envío de la sub-tarea hasta HOY (solo sustis activas en cache).</div>')

    # === Tabla detalle compartida ===
    JIRA_URL = "https://leaseir.atlassian.net/browse/"
    COLS = [
        ("chain", "Cadena", "text", True),
        ("k", "Ticket", "text", True),
        ("cliente", "Cliente", "text", True),
        ("status", "Estado actual", "text", True),
        ("tipo", "Tipo avería", "text", True),
        ("motivo", "Resumen avería", "text", False),
        ("cambios", "Cambios/Mejoras", "text", False),
        ("bloq", "Bloq", "text", True),
        ("susti", "Susti", "text", True),
        ("susti_status", "Estado susti", "text", True),
        ("dias_taller", "Días taller", "num", True),
        ("dias_susti", "Días susti", "num", True),
        ("consola", "Consola", "text", False),
        ("hp", "HP", "text", False),
        ("garantia", "Garantía", "text", False),
        ("disparos", "Disparos", "num", False),
        ("importe", "Importe", "num", False),
        ("fventa", "Fecha venta", "date", False),
        ("entrada_taller", "Entrada taller", "date", False),
        ("salida_taller", "Salida taller", "date", False),
        ("created", "Creada", "date", True),
        ("asignado", "Asignado", "text", False),
        ("tec_taller", "Téc. taller", "text", False),
        ("tec_externo", "Téc. externo", "text", False),
    ]

    chains_set = sorted({r["chain"] for r in rows})
    statuses_set = sorted({r["status"] for r in rows if r["status"]})
    tipos_set = sorted({r["tipo"] for r in rows if r["tipo"]})

    toolbar = ['<div class="toolbar" style="flex-wrap:wrap;gap:6px">']
    toolbar.append('<input type="text" id="tm-search" placeholder="Buscar texto...">')
    toolbar.append('<select id="tm-chain"><option value="">Cadena: todas</option>' + ''.join(f'<option>{html_escape(c)}</option>' for c in chains_set) + '</select>')
    toolbar.append('<select id="tm-status"><option value="">Estado: todos</option>' + ''.join(f'<option>{html_escape(s)}</option>' for s in statuses_set) + '</select>')
    toolbar.append('<select id="tm-tipo"><option value="">Tipo: todos</option>' + ''.join(f'<option>{html_escape(s)}</option>' for s in tipos_set) + '</select>')
    toolbar.append('<select id="tm-bloq"><option value="">Bloq: todos</option><option>Sí</option><option>No</option></select>')
    toolbar.append('<select id="tm-susti"><option value="">Susti: todas</option><option>Sí</option><option>No</option></select>')
    toolbar.append('<select id="tm-open"><option value="">Estado: todos</option><option value="open">Abiertas (live)</option><option value="closed">Cerradas</option></select>')
    toolbar.append('<button type="button" id="tm-clear" class="multi-btn">✕ Limpiar</button>')
    toolbar.append('<span class="count"><b id="tm-count">0</b> de <span id="tm-total">0</span> visibles</span>')
    toolbar.append('</div>')
    # Banner KPI activo
    toolbar.append('<div id="tm-kpi-banner" style="display:none;background:#e7eefa;border:1px solid #2a59c4;border-radius:6px;padding:6px 12px;margin:6px 0;font-size:13px;color:#1f3a5f"><i class="ti ti-filter"></i> Filtro KPI activo: <b class="tm-banner-label">—</b> <button type="button" id="tm-kpi-clear" style="margin-left:10px;border:1px solid #2a59c4;background:white;border-radius:4px;padding:2px 8px;font-size:12px;cursor:pointer">✕ Quitar</button></div>')

    # Toggle columnas
    toggle = ['<details style="margin:4px 0 8px"><summary style="cursor:pointer;color:var(--blue);font-size:12px">Mostrar/ocultar columnas</summary>',
              '<div style="display:flex;flex-wrap:wrap;gap:8px;padding:8px;background:#f8fafc;border:1px solid var(--line);border-radius:6px;font-size:11px">']
    for i, (key, label, dtype, default_on) in enumerate(COLS):
        checked = " checked" if default_on else ""
        toggle.append(f'<label><input type="checkbox" data-col-idx="{i}"{checked}> {html_escape(label)}</label>')
    toggle.append('</div></details>')

    # Header + rows
    out = []
    out.append(''.join(toolbar))
    out.append(''.join(toggle))
    out.append('<div class="scroller"><table id="tm-table" style="font-size:11px;min-width:2200px"><thead><tr>')
    for i, (key, label, dtype, default_on) in enumerate(COLS):
        hide = "" if default_on else 'style="display:none"'
        out.append(f'<th class="sortable tm-col" data-col="{i}" data-type="{dtype}" {hide}>{html_escape(label)}</th>')
    out.append('</tr></thead><tbody>')
    # Sort por días taller desc
    rows.sort(key=lambda r: -(r["dias_taller"] if r["dias_taller"] is not None else -1))
    for r in rows:
        link = f'<a href="{JIRA_URL}{r["k"]}" target="_blank" rel="noopener" style="color:#2a59c4">{r["k"]}</a>'
        cell_vals = []
        for key, label, dtype, default_on in COLS:
            v = r.get(key, "")
            if key == "k":
                v = link
            elif key == "dias_taller":
                if r["taller_live"]:
                    v = f'<span title="Aún en taller hoy" style="color:#c0392b;font-weight:600">{v}*</span>' if v is not None else "—"
                else:
                    v = v if v is not None else "—"
            elif key == "dias_susti":
                v = v if v is not None else "—"
            elif key == "disparos":
                v = format_int_dot(v) if v != "" else ""
            else:
                v = html_escape(str(v)) if v is not None else ""
            hide = "" if default_on else 'style="display:none"'
            cell_vals.append(f'<td class="tm-col" data-col="{COLS.index((key, label, dtype, default_on))}" {hide}>{v}</td>')
        row_open = "open" if r["is_open"] else "closed"
        out.append(f'<tr data-open="{row_open}" data-chain="{html_escape(r["chain"])}" data-status="{html_escape(r["status"])}" data-tipo="{html_escape(r["tipo"])}" data-bloq="{html_escape(r["bloq"])}" data-susti="{html_escape(r["susti"])}" data-dt="{r["dias_taller"] if r["dias_taller"] is not None else -1}" data-ds="{r["dias_susti"] if r["dias_susti"] is not None else -1}">' + ''.join(cell_vals) + '</tr>')
    out.append('</tbody></table></div>')

    # JS para tabla
    out.append('''<script>(function(){
      var tbl=document.getElementById('tm-table');if(!tbl)return;
      var rows=Array.from(tbl.tBodies[0].rows);
      var cnt=document.getElementById('tm-count');document.getElementById('tm-total').textContent=rows.length;
      var fs={search:'tm-search',chain:'tm-chain',status:'tm-status',tipo:'tm-tipo',bloq:'tm-bloq',susti:'tm-susti',open:'tm-open'};
      var els={};for(var k in fs)els[k]=document.getElementById(fs[k]);
      // Filtros KPI Tiempos: {kpi: 'taller'|'sustis', mode: 'all'|'top10'|'bucket', bucket: [min,max]}
      var kpiFilter = null;
      function apply(){
        var q=(els.search.value||'').toLowerCase().trim();
        var n=0;
        // Para top10, calcular threshold del campo correcto
        var fieldKey = kpiFilter && kpiFilter.kpi === 'sustis' ? 'ds' : 'dt';
        var sortedVs=[];
        rows.forEach(function(r){ var d=parseInt(r.dataset[fieldKey])||0; if(d>=0)sortedVs.push(d); });
        sortedVs.sort(function(a,b){return b-a;});
        var threshold = (kpiFilter && kpiFilter.mode==='top10') ? sortedVs[Math.max(0,Math.floor(sortedVs.length*0.1)-1)] : -1;
        for(var i=0;i<rows.length;i++){var r=rows[i];
          var dt = parseInt(r.dataset.dt); var ds = parseInt(r.dataset.ds);
          var v = (kpiFilter && kpiFilter.kpi === 'sustis') ? ds : dt;
          var kpiOk = true;
          if (kpiFilter) {
            // 'all': solo filas con tiempo medido (>=0)
            if (kpiFilter.mode === 'all') kpiOk = (v >= 0);
            else if (kpiFilter.mode === 'top10') kpiOk = (v >= 0 && v >= threshold);
            else if (kpiFilter.mode === 'bucket') kpiOk = (v >= kpiFilter.bucket[0] && v <= kpiFilter.bucket[1]);
          }
          var ok=(!q||r.textContent.toLowerCase().indexOf(q)>=0)
            && (!els.chain.value||r.dataset.chain===els.chain.value)
            && (!els.status.value||r.dataset.status===els.status.value)
            && (!els.tipo.value||r.dataset.tipo===els.tipo.value)
            && (!els.bloq.value||r.dataset.bloq===els.bloq.value)
            && (!els.susti.value||r.dataset.susti===els.susti.value)
            && (!els.open.value||r.dataset.open===els.open.value)
            && kpiOk;
          r.style.display=ok?'':'none';if(ok)n++;
        }
        cnt.textContent=n;
        // Mostrar banner del filtro KPI activo
        var banner = document.getElementById('tm-kpi-banner');
        if (banner) {
          if (kpiFilter) {
            var lbl = kpiFilter.kpi === 'sustis' ? 'Sustituciones' : 'Tiempo en taller';
            var modeLbl = kpiFilter.mode === 'all' ? ' (todos con tiempo medido)' :
                          kpiFilter.mode === 'top10' ? ' (top 10% mayor tiempo)' :
                          kpiFilter.mode === 'bucket' ? (' (bucket ' + kpiFilter.bucket[0] + '-' + (kpiFilter.bucket[1]===9999?'∞':kpiFilter.bucket[1]) + 'd)') : '';
            banner.style.display = '';
            banner.querySelector('.tm-banner-label').textContent = lbl + modeLbl;
          } else {
            banner.style.display = 'none';
          }
        }
      }
      Object.values(els).forEach(function(e){e.addEventListener('input',apply);e.addEventListener('change',apply);});
      document.getElementById('tm-clear').addEventListener('click',function(){Object.values(els).forEach(function(e){e.value='';});kpiFilter=null;apply();});
      var bannerClose = document.getElementById('tm-kpi-clear');
      if (bannerClose) bannerClose.addEventListener('click', function(){ kpiFilter = null; apply(); });
      // Toggle columnas
      document.querySelectorAll('input[data-col-idx]').forEach(function(cb){
        cb.addEventListener('change',function(){
          var i=cb.dataset.colIdx;
          document.querySelectorAll('.tm-col[data-col="'+i+'"]').forEach(function(el){
            el.style.display = cb.checked ? '' : 'none';
          });
        });
      });
      // Sort por columna
      var heads=tbl.tHead.rows[0].cells;var lastSort={col:-1,dir:1};
      for(var hi=0;hi<heads.length;hi++){(function(th){th.style.cursor='pointer';th.addEventListener('click',function(){
        var col=parseInt(th.dataset.col);var type=th.dataset.type||'text';var dir=(lastSort.col===col)?-lastSort.dir:1;lastSort={col:col,dir:dir};
        var rs=Array.from(tbl.tBodies[0].rows);
        rs.sort(function(a,b){var av=a.cells[col].textContent.trim();var bv=b.cells[col].textContent.trim();
          if(type==='num'){av=parseFloat(av.replace(/\./g,'').replace(',','.'))||0;bv=parseFloat(bv.replace(/\./g,'').replace(',','.'))||0;return(av-bv)*dir;}
          return av.localeCompare(bv,'es')*dir;});
        for(var k=0;k<rs.length;k++)tbl.tBodies[0].appendChild(rs[k]);
      });})(heads[hi]);}
      apply();
      // KPIs Tiempos clickables — filtran tabla
      document.querySelectorAll('[data-tiempos-filter]').forEach(function(kpi){
        kpi.addEventListener('click',function(){
          var f = kpi.dataset.tiemposFilter;
          var k = kpi.dataset.tiemposKpi || 'taller';
          if (f === 'all') kpiFilter = { kpi: k, mode: 'all' };
          else if (f === 'max') kpiFilter = { kpi: k, mode: 'top10' };
          else kpiFilter = { kpi: k, mode: 'all' };  // mediana y media también filtran
          Object.values(els).forEach(function(e){e.value='';});
          apply();
          tbl.scrollIntoView({behavior:'smooth',block:'start'});
        });
      });
      // Stacked bar segmentos clickables — filtran por bucket de días
      document.querySelectorAll('[data-tiempos-bucket]').forEach(function(seg){
        seg.style.cursor = 'pointer';
        seg.addEventListener('click', function(){
          var k = seg.dataset.tiemposKpi || 'taller';
          var bucket = seg.dataset.tiemposBucket.split(',').map(function(x){return parseInt(x);});
          kpiFilter = { kpi: k, mode: 'bucket', bucket: bucket };
          Object.values(els).forEach(function(e){e.value='';});
          apply();
          tbl.scrollIntoView({behavior:'smooth',block:'start'});
        });
      });
    })();</script>''')

    detail_html = ''.join(out)

    return {"taller": taller_html, "sustis": sustis_html, "detail": detail_html}
# ========== FIN TIEMPOS REFACTORIZADOS ==========




def build_tecnicos_dia_section(cache):
    """Tabla agregada técnicos x mes con #presupuestos + #reparaciones desde 1-ene-2026.
    NOTA: la atribución es heurística (tec_taller ACTUAL del ticket), porque no
    tenemos changelog del campo. Funciona razonablemente al agregar por mes."""
    PRES_STATE = "Presupuesto preparado pendiente de enviar"
    REP_STATE = "Inspección de salida"
    cutoff = datetime(2026, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    MONTH_LABELS = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

    # grid[(tec, month_num)] = {"p": N, "r": N}
    grid = {}
    tecs = set()
    months_seen = set()
    for k, t in cache.get("tickets", {}).items():
        tec = (t.get("tec_taller") or "").strip() or (t.get("asignado") or "").strip() or "(Sin asignar)"
        for tr in t.get("transitions", []):
            if len(tr) < 3: continue
            ts, fs, to = tr[0], tr[1], tr[2]
            if to not in (PRES_STATE, REP_STATE): continue
            dt = parse_iso(ts)
            if not dt or dt < cutoff: continue
            m = dt.month
            months_seen.add(m)
            tecs.add(tec)
            cell = grid.setdefault((tec, m), {"p": 0, "r": 0})
            if to == PRES_STATE: cell["p"] += 1
            else: cell["r"] += 1

    if not grid:
        return '<div style="color:#94a3b8;padding:14px;font-style:italic">Sin datos desde 1-ene-2026.</div>'

    months_sorted = sorted(months_seen)
    tecs_sorted = sorted(tecs, key=lambda x: (0 if x == "(Sin asignar)" else 1, x.lower()))

    out = []
    out.append('<div class="legend">Vista agregada: filas = técnicos, columnas = meses 2026. Cada celda muestra <b style="color:#0891b2">P presupuestos</b> + <b style="color:#059669">R reparaciones</b>. <i>Atribución heurística por tec_taller actual.</i></div>')
    out.append('<div class="scroller"><table style="font-size:12px;min-width:800px">')
    out.append('<thead><tr>')
    out.append('<th class="sticky-l1" style="text-align:left;left:0;min-width:180px;background:#f1f5f9;color:#0b3d91">Técnico</th>')
    for m in months_sorted:
        out.append(f'<th style="background:#f1f5f9;font-size:12px;padding:6px 10px;text-align:center">{MONTH_LABELS[m-1]}</th>')
    out.append('<th style="background:#fff4d6;font-size:12px;padding:6px 12px;text-align:center">Total YTD</th>')
    out.append('</tr></thead><tbody>')

    # Totales fila
    col_tot = {m: {"p": 0, "r": 0} for m in months_sorted}
    grand = {"p": 0, "r": 0}
    for tec in tecs_sorted:
        row_p, row_r = 0, 0
        cells = [f'<td class="lbl-name" style="left:0;min-width:180px">{html_escape(tec)}</td>']
        for m in months_sorted:
            v = grid.get((tec, m))
            if v:
                p, r = v["p"], v["r"]
                row_p += p; row_r += r
                col_tot[m]["p"] += p; col_tot[m]["r"] += r
                parts = []
                if p: parts.append(f'<b style="color:#0891b2">{p}</b>')
                if r: parts.append(f'<b style="color:#059669">{r}</b>')
                inner = "+".join(parts) if parts else "·"
                cells.append(f'<td class="tec-mes-cell" data-tec="{html_escape(tec)}" data-mes="{m}" style="text-align:center;cursor:pointer" title="{p} ppto · {r} rep">{inner}</td>')
            else:
                cells.append('<td style="text-align:center;color:#cbd5e1">·</td>')
        if row_p or row_r:
            cells.append(f'<td class="tec-mes-cell" data-tec="{html_escape(tec)}" data-mes="all" style="background:#fff4d6;text-align:center;font-weight:600;cursor:pointer"><span style="color:#0891b2">{row_p}P</span>+<span style="color:#059669">{row_r}R</span></td>')
            grand["p"] += row_p; grand["r"] += row_r
        else:
            cells.append('<td style="background:#fff4d6;text-align:center;color:#cbd5e1">·</td>')
        out.append('<tr>' + ''.join(cells) + '</tr>')
    # Fila total
    cells = ['<td class="lbl-name" style="left:0;min-width:180px;background:#fff4d6;font-weight:600">Total mes</td>']
    for m in months_sorted:
        p = col_tot[m]["p"]; r = col_tot[m]["r"]
        cells.append(f'<td style="background:#fff4d6;text-align:center;font-weight:600"><span style="color:#0891b2">{p}P</span>+<span style="color:#059669">{r}R</span></td>')
    cells.append(f'<td style="background:#fff4d6;text-align:center;font-weight:700"><span style="color:#0891b2">{grand["p"]}P</span>+<span style="color:#059669">{grand["r"]}R</span></td>')
    out.append('<tr style="background:#fff4d6">' + ''.join(cells) + '</tr>')
    out.append('</tbody></table></div>')
    return ''.join(out)



def build_tiempo_taller_section(cache):
    """Stacked bar de tiempo en taller para tickets 2026 cerrados."""
    TALLER = {"Pendiente asignar técnico", "En cola taller",
              "En preparación presupuesto", "Esperando inicio reparación",
              "En reparación", "Inspección de salida"}
    EXIT = {"Devuelto a cliente", "Resuelto", "Finalizada", "Cancelado",
            "Finalizado técnico externo"}
    BUCKETS = [
        ("<4 días",   "#1f8a4c", lambda d: d < 4),
        ("5-8 días",  "#65a30d", lambda d: 4 <= d <= 8),
        ("9-15 días", "#b8860b", lambda d: 9 <= d <= 15),
        ("16-30 días","#cb6f0a", lambda d: 16 <= d <= 30),
        (">30 días",  "#c0392b", lambda d: d > 30),
    ]
    counts = [0] * len(BUCKETS)
    dias_list = []
    sample_max = []
    for k, t in cache.get("tickets", {}).items():
        if is_open(t.get("current_status")):
            continue
        if not (t.get("created") or "").startswith("2026"):
            continue
        entrada = None
        salida = None
        for tr in t.get("transitions", []):
            if len(tr) < 3: continue
            ts, fs, to = tr[0], tr[1], tr[2]
            dt = parse_iso(ts)
            if not dt: continue
            if entrada is None and to == "Pendiente asignar técnico":
                entrada = dt
            if entrada and (to in EXIT) and (fs in TALLER):
                salida = dt
                break
        if entrada and salida:
            dias = (salida - entrada).days
            if dias < 0: continue
            dias_list.append(dias)
            sample_max.append((dias, k))
            for i, (_, _, pred) in enumerate(BUCKETS):
                if pred(dias):
                    counts[i] += 1
                    break
    total = sum(counts)
    if not total:
        return '<div style="color:#94a3b8;padding:14px;font-style:italic">Aún no hay tickets cerrados con tiempo en taller calculable.</div>'
    media = sum(dias_list) / total
    mediana = sorted(dias_list)[total // 2]
    max_d = max(dias_list)
    sample_max.sort(reverse=True)
    top3 = sample_max[:3]
    segments = []
    for (lbl, color, _), n in zip(BUCKETS, counts):
        pct = (n * 100.0 / total) if total else 0
        if n > 0:
            label_inline = "" if pct < 5 else str(n)
            segments.append(
                f'<div style="background:{color};height:100%;width:{pct:.2f}%;display:flex;align-items:center;justify-content:center;color:white;font-weight:600;font-size:13px;min-width:0;overflow:hidden" title="{lbl}: {n} tickets ({pct:.0f}%)">{label_inline}</div>'
            )
    bar_html = '<div style="display:flex;height:36px;border-radius:6px;overflow:hidden;border:1px solid var(--line);background:#f1f5f9">' + "".join(segments) + '</div>'
    legend_items = []
    for (lbl, color, _), n in zip(BUCKETS, counts):
        pct = (n * 100.0 / total) if total else 0
        legend_items.append(
            f'<span style="display:inline-flex;align-items:center;gap:6px;margin-right:14px;font-size:13px">'
            f'<span style="display:inline-block;width:12px;height:12px;background:{color};border-radius:3px"></span>'
            f'<b>{lbl}</b>: {n} ({pct:.0f}%)</span>'
        )
    top_label = html_escape(top3[0][1]) if top3 else ""
    kpis = (
        '<div style="display:flex;gap:12px;margin:6px 0 12px;flex-wrap:wrap">'
        f'<div class="kpi" style="flex:1 1 150px"><div class="k">Total tickets</div><div class="v">{total}</div><div class="d">cerrados 2026 con tiempo medible</div></div>'
        f'<div class="kpi" style="flex:1 1 150px"><div class="k">Mediana</div><div class="v">{mediana:.0f}<span style="font-size:14px;color:var(--grey);margin-left:4px">días</span></div><div class="d">la mitad se cierran en este plazo</div></div>'
        f'<div class="kpi" style="flex:1 1 150px"><div class="k">Media</div><div class="v">{media:.1f}<span style="font-size:14px;color:var(--grey);margin-left:4px">días</span></div><div class="d">sensible a outliers</div></div>'
        f'<div class="kpi" style="flex:1 1 150px"><div class="k">Máximo</div><div class="v" style="color:#c0392b">{max_d}<span style="font-size:14px;color:var(--grey);margin-left:4px">días</span></div><div class="d">{top_label}</div></div>'
        '</div>'
    )
    return (
        kpis + bar_html +
        '<div style="margin-top:12px">' + "".join(legend_items) + '</div>'
        '<div class="legend" style="margin-top:8px">Tiempo = primera entrada a <b>Pdte. asignar técnico</b> hasta primera salida a estado terminal (Devuelto/Resuelto/Finalizada/Cancelado).</div>'
    )


FONT_X = "Arial"
HDR_BLUE = "0B3D91"
ZEBRA = "F5F7FB"
TOTAL_BG = "FFF4D6"
FUNNEL_C_BG = "F0EAFB"
thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr_cell(c, t):
    c.value = t
    c.font = Font(name=FONT_X, bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", start_color=HDR_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


def lbl(c, t, bold=False, fill=None):
    c.value = t
    c.font = Font(name=FONT_X, bold=bold, size=10)
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = BORDER


def num(c, v, bold=False, fill=None):
    c.value = v
    c.font = Font(name=FONT_X, bold=bold, size=10)
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER


def build_excel(cache, out_path):
    today = date.today()
    cutoffs = compute_cutoffs(today)
    opens = replay_opens(cache, cutoffs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Evol_Estado"
    ws["A1"] = f"Evolución abiertas SAT por estado — {today.isoformat()}"
    ws["A1"].font = Font(name=FONT_X, bold=True, size=13, color=HDR_BLUE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cutoffs) + 2)

    hdr_cell(ws.cell(row=3, column=1), "Funnel")
    hdr_cell(ws.cell(row=3, column=2), "Estado")
    for j, (l, _) in enumerate(cutoffs, 3):
        hdr_cell(ws.cell(row=3, column=j), l)

    present = set()
    for l, _ in cutoffs:
        for _, _, s in opens[l]:
            if s:
                present.add(s)
    states = [s for s in FUNNEL_ORDER if s in present] + sorted(s for s in present if s not in FUNNEL_ORDER)

    row = 4
    for st in states:
        ft = FUNNEL_TAG.get(st, "?")
        f = FUNNEL_C_BG if ft == "C" else (ZEBRA if (row % 2 == 0) else None)
        lbl(ws.cell(row=row, column=1), ft, fill=f)
        lbl(ws.cell(row=row, column=2), st, fill=f)
        for j, (l, _) in enumerate(cutoffs, 3):
            c = sum(1 for _, _, s in opens[l] if s == st)
            num(ws.cell(row=row, column=j), c if c else "", fill=f)
        row += 1
    lbl(ws.cell(row=row, column=1), "", bold=True, fill=TOTAL_BG)
    lbl(ws.cell(row=row, column=2), "Total abiertas", bold=True, fill=TOTAL_BG)
    for j, (l, _) in enumerate(cutoffs, 3):
        num(ws.cell(row=row, column=j), len(opens[l]), bold=True, fill=TOTAL_BG)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 38
    for j in range(len(cutoffs)):
        ws.column_dimensions[get_column_letter(3 + j)].width = 12
    ws.freeze_panes = "C4"

    ws2 = wb.create_sheet("Evol_Cadena")
    ws2["A1"] = f"Evolución abiertas SAT por cadena — {today.isoformat()}"
    ws2["A1"].font = Font(name=FONT_X, bold=True, size=13, color=HDR_BLUE)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cutoffs) + 1)
    hdr_cell(ws2.cell(row=3, column=1), "Cadena")
    for j, (l, _) in enumerate(cutoffs, 2):
        hdr_cell(ws2.cell(row=3, column=j), l)
    row = 4
    for ch in CHAIN_ORDER:
        f = "EAEAEA" if ch == "Otros" else (ZEBRA if (row % 2 == 0) else None)
        lbl(ws2.cell(row=row, column=1), ch, fill=f, bold=(ch == "Otros"))
        for j, (l, _) in enumerate(cutoffs, 2):
            c = sum(1 for _, x, _ in opens[l] if x == ch)
            num(ws2.cell(row=row, column=j), c if c else "", fill=f)
        row += 1
    lbl(ws2.cell(row=row, column=1), "Total", bold=True, fill=TOTAL_BG)
    for j, (l, _) in enumerate(cutoffs, 2):
        num(ws2.cell(row=row, column=j), len(opens[l]), bold=True, fill=TOTAL_BG)
    ws2.column_dimensions["A"].width = 18
    for j in range(len(cutoffs)):
        ws2.column_dimensions[get_column_letter(2 + j)].width = 12
    ws2.freeze_panes = "B4"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def html_escape(s):
    if s is None:
        return ""
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


TALLER_STATUSES_ORDER = [
    "Pendiente asignar técnico", "En cola taller", "En preparación presupuesto",
    "Presupuesto preparado pendiente de enviar", "Pendiente confirmación presupuesto",
    "Esperando inicio reparación", "En reparación", "Inspección de salida",
]


def build_abiertas_hoy_section(cache):
    rows_data = []
    for key, t in cache.get("tickets", {}).items():
        if not is_open(t.get("current_status")):
            continue
        chain = chain_of(t.get("cliente", ""), t.get("loc", ""))
        days = days_since(last_status_change_dt(t))
        created_today = is_created_today(t)
        rows_data.append((chain, t.get("current_status"), days, created_today))

    grid = {}
    for ch, st, d, ct in rows_data:
        k = (st, ch)
        if k not in grid:
            grid[k] = {"t": 0, "g": 0, "y": 0, "r": 0, "total": 0}
        # Azul (creado hoy) tiene prioridad sobre verde/amarillo/rojo
        bucket = "t" if ct else days_bucket(d)
        grid[k][bucket] += 1
        grid[k]["total"] += 1

    changes_today = compute_status_changes_today(cache)

    present = sorted({k[0] for k in grid.keys()})
    states_order = [s for s in FUNNEL_ORDER if s in present]
    states_order += sorted(s for s in present if s not in FUNNEL_ORDER)
    taller_set = set(TALLER_STATUSES_ORDER)

    COLOR_TODAY = "#1e6fdb"
    ARROW_IN = "#c0392b"
    ARROW_OUT = "#1f8a4c"

    def cell_rich(d, ch_in=0, ch_out=0):
        if d["total"] == 0 and ch_in == 0 and ch_out == 0:
            return '<span style="color:#c0d0e0">·</span>'
        if d["total"] == 0:
            body = '<span style="color:#c0d0e0">·</span>'
        else:
            inner = ""
            if d["t"] > 0:
                inner += f'<span style="color:{COLOR_TODAY};font-weight:500">{d["t"]}</span><span style="font-size:11px;color:#7d8590">/</span>'
            inner += (f'<span style="color:{COLOR_GREEN};font-weight:500">{d["g"]}</span>'
                      f'<span style="font-size:11px;color:#7d8590">/</span>'
                      f'<span style="color:{COLOR_YELLOW};font-weight:500">{d["y"]}</span>'
                      f'<span style="font-size:11px;color:#7d8590">/</span>'
                      f'<span style="color:{COLOR_RED};font-weight:500">{d["r"]}</span>')
            body = (f'<span style="font-weight:500">{d["total"]}</span> '
                    f'<span style="font-size:11px;color:#7d8590">(</span>'
                    f'{inner}'
                    f'<span style="font-size:11px;color:#7d8590">)</span>')
        flow = []
        if ch_in > 0:
            flow.append(f'<span style="color:{ARROW_IN};font-weight:500;font-size:11px" title="entran hoy">&#9650;{ch_in}</span>')
        if ch_out > 0:
            flow.append(f'<span style="color:{ARROW_OUT};font-weight:500;font-size:11px" title="salen hoy">&#9660;{ch_out}</span>')
        if flow:
            body += ' <span style="font-size:11px;color:#7d8590">·</span> ' + ' '.join(flow)
        return body

    parts = ['<tr><th class="sticky-l1">Funnel</th><th class="sticky-l2">Estado</th>']
    for ch in CHAIN_ORDER:
        parts.append(f'<th>{ch}</th>')
    parts.append('<th class="col-total">Total</th></tr>')
    header = "".join(parts)

    body_parts = []
    chain_tot = {ch: {"t": 0, "g": 0, "y": 0, "r": 0, "total": 0} for ch in CHAIN_ORDER}
    chain_flow = {ch: {"in": 0, "out": 0} for ch in CHAIN_ORDER}
    grand = {"t": 0, "g": 0, "y": 0, "r": 0, "total": 0}
    grand_flow = {"in": 0, "out": 0}
    for st in states_order:
        ft = FUNNEL_TAG.get(st, "?")
        color = FUNNEL_COLOR.get(ft, "#888")
        row_tot = {"t": 0, "g": 0, "y": 0, "r": 0, "total": 0}
        row_flow = {"in": 0, "out": 0}
        is_taller = st in taller_set
        idx = states_order.index(st)
        prev_taller = (idx > 0 and states_order[idx-1] in taller_set)
        next_taller = (idx < len(states_order)-1 and states_order[idx+1] in taller_set)
        row_classes = []
        if is_taller:
            row_classes.append("row-taller")
            if not prev_taller:
                row_classes.append("row-taller-first")
            if not next_taller:
                row_classes.append("row-taller-last")
        row_cls_attr = f' class="{" ".join(row_classes)}"' if row_classes else ""
        cells = [f'<tr{row_cls_attr}><td class="lbl-funnel" style="color:{color}">{ft}</td>',
                 f'<td class="lbl-name">{html_escape(st)}</td>']
        for ch in CHAIN_ORDER:
            d = grid.get((st, ch), {"t": 0, "g": 0, "y": 0, "r": 0, "total": 0})
            ch_changes = changes_today.get((st, ch), {"in": 0, "out": 0})
            clickable = (d["total"] > 0)
            cls = "num ah-cell" if clickable else "num"
            attrs = f' data-estado="{html_escape(st)}" data-cadena="{html_escape(ch)}"' if clickable else ""
            cells.append(f'<td class="{cls}"{attrs}>{cell_rich(d, ch_changes["in"], ch_changes["out"])}</td>')
            for k in row_tot:
                row_tot[k] += d[k]
                chain_tot[ch][k] += d[k]
                grand[k] += d[k]
            row_flow["in"] += ch_changes["in"]
            row_flow["out"] += ch_changes["out"]
            chain_flow[ch]["in"] += ch_changes["in"]
            chain_flow[ch]["out"] += ch_changes["out"]
            grand_flow["in"] += ch_changes["in"]
            grand_flow["out"] += ch_changes["out"]
        cls_tot = "num col-total ah-cell" if row_tot["total"] > 0 else "num col-total"
        attrs_tot = f' data-estado="{html_escape(st)}" data-cadena=""' if row_tot["total"] > 0 else ""
        cells.append(f'<td class="{cls_tot}"{attrs_tot}>{cell_rich(row_tot, row_flow["in"], row_flow["out"])}</td>')
        cells.append('</tr>')
        body_parts.append("".join(cells))
    tot_cells = ['<tr class="total"><td class="lbl-funnel"></td><td class="lbl-name">Total</td>']
    for ch in CHAIN_ORDER:
        d = chain_tot[ch]
        cf = chain_flow[ch]
        clickable = (d["total"] > 0)
        cls = "num ah-cell" if clickable else "num"
        attrs = f' data-estado="" data-cadena="{html_escape(ch)}"' if clickable else ""
        tot_cells.append(f'<td class="{cls}"{attrs}>{cell_rich(d, cf["in"], cf["out"])}</td>')
    g = grand
    cls_g = "num col-total ah-cell" if g["total"] > 0 else "num col-total"
    attrs_g = ' data-estado="" data-cadena=""' if g["total"] > 0 else ""
    tot_cells.append(f'<td class="{cls_g}"{attrs_g}>{cell_rich(g, grand_flow["in"], grand_flow["out"])}</td>')
    tot_cells.append('</tr>')
    body_parts.append("".join(tot_cells))

    return header, "\n".join(body_parts)


def build_detalle_section(cache, sustis_by_parent=None):
    sustis_by_parent = sustis_by_parent or {}
    JIRA_URL = "https://leaseir.atlassian.net/browse/"
    rows = []
    for key, t in cache.get("tickets", {}).items():
        if not is_open(t.get("current_status")):
            continue
        chain = chain_of(t.get("cliente", ""), t.get("loc", ""))
        days = days_since(last_status_change_dt(t))
        created_dt = parse_iso(t.get("created"))
        created_s = created_dt.strftime("%d/%m/%Y") if created_dt else ""
        fventa_dt = parse_iso(t.get("fventa") or "")
        fventa_s = fventa_dt.strftime("%d/%m/%Y") if fventa_dt else (t.get("fventa") or "")
        st = t.get("current_status") or ""
        ft = FUNNEL_TAG.get(st, "?")
        ft_color = FUNNEL_COLOR.get(ft, "#888")
        days_color = COLOR_GREEN if (days is not None and days < 5) else (
            COLOR_YELLOW if (days is not None and days <= 15) else COLOR_RED)
        dias_taller = None
        dt_taller = None
        for tr in t.get("transitions", []):
            if len(tr) >= 3 and tr[2] == "Pendiente asignar técnico":
                _dt = parse_iso(tr[0])
                if _dt:
                    dt_taller = _dt
                    dias_taller = days_since(_dt)
                    break
        dt_taller_label = dt_taller.strftime("%d/%m/%Y") if dt_taller else ""
        dt_color = COLOR_GREEN if (dias_taller is not None and dias_taller < 5) else (
            COLOR_YELLOW if (dias_taller is not None and dias_taller <= 15) else COLOR_RED)
        rows.append({
            "chain": chain, "key": key,
            "cliente": t.get("cliente", ""),
            "status": st, "ft": ft, "ft_color": ft_color,
            "days": days if days is not None else "",
            "days_color": days_color, "loc": t.get("loc", ""),
            "dias_taller": dias_taller if dias_taller is not None else "",
            "dt_taller_label": dt_taller_label,
            "dt_color": dt_color if dias_taller is not None else "#cbd5e1",
            "created": created_s, "tipo": t.get("tipo", ""),
            "bloq": t.get("bloq", ""), "susti": t.get("susti", ""),
            "susti_status": sustis_by_parent.get(key, ""),
            "fventa": fventa_s, "consola": t.get("consola", ""),
            "hp": t.get("hp", ""), "garantia": t.get("garantia", ""),
            "disparos": t.get("disparos", ""),
            "desc": (t.get("descripcion", "") or "")[:200],
            "tec_taller": t.get("tec_taller", ""),
            "asignado": t.get("asignado", ""),
            "importe": t.get("importe", ""),
            "tec_externo": t.get("tec_externo", ""),
            "ult_com": (t.get("ult_comentario", "") or "")[:300],
        })

    idx = {s: i for i, s in enumerate(FUNNEL_ORDER)}
    rows.sort(key=lambda r: (CHAIN_ORDER.index(r["chain"]) if r["chain"] in CHAIN_ORDER else 99,
                              idx.get(r["status"], 999),
                              -(r["days"] if isinstance(r["days"], int) else 0)))

    body = []
    for r in rows:
        link = f'<a href="{JIRA_URL}{r["key"]}" target="_blank" rel="noopener" style="color:#2a59c4;text-decoration:none">{r["key"]}</a>'
        body.append(
            ('<tr data-no-tec="1">' if r["status"] in {"Pendiente asignar técnico","En cola taller","Presupuesto preparado pendiente de enviar","Pendiente confirmación presupuesto","Inspección de salida"} else '<tr>')
            + f'<td>{html_escape(r["chain"])}</td>'
            f'<td>{link}</td>'
            f'<td title="{html_escape(r["cliente"])}" class="d-trunc">{html_escape(r["cliente"])}</td>'
            f'<td>{html_escape(r["status"])}</td>'
            f'<td style="color:{r["ft_color"]};font-weight:500;text-align:center">{r["ft"]}</td>'
            f'<td style="color:{r["days_color"]};font-weight:500;text-align:center">{r["days"]}</td>'
            f'<td style="color:{r["dt_color"]};font-weight:500;text-align:center" title="Entró en Pdte. asignar técnico el {r["dt_taller_label"] or "—"}">{r["dias_taller"] if r["dias_taller"] != "" else "—"}</td>'
            f'<td title="{html_escape(r["loc"])}" class="d-trunc">{html_escape(r["loc"])}</td>'
            f'<td>{r["created"]}</td>'
            f'<td>{html_escape(r["tipo"])}</td>'
            f'<td style="text-align:center">{html_escape(r["bloq"])}</td>'
            f'<td style="text-align:center">{html_escape(r["susti"])}</td>'
            f'<td style="text-align:center;font-size:11px">{html_escape(r["susti_status"]) if r["susti_status"] else "—"}</td>'
            f'<td>{r["fventa"]}</td>'
            f'<td style="text-align:center">{html_escape(r["consola"])}</td>'
            f'<td style="text-align:center">{html_escape(r["hp"])}</td>'
            f'<td style="text-align:right;font-variant-numeric:tabular-nums">{html_escape(format_int_dot(r["disparos"]))}</td>'
            f'<td style="text-align:center">{html_escape(r["garantia"])}</td>'
            f'<td title="{html_escape(r["desc"])}" class="d-trunc">{html_escape(r["desc"])}</td>'
            f'<td>{html_escape(r["tec_taller"])}</td>'
            f'<td style="text-align:right">{html_escape(r["importe"])}</td>'
            f'<td>{html_escape(r["tec_externo"])}</td>'
            f'<td title="{html_escape(r["ult_com"])}" class="d-trunc">{html_escape(r["ult_com"])}</td>'
            f'<td>{html_escape(r["asignado"])}</td>'
            '</tr>'
        )
    return "\n".join(body)


def build_html(cache, out_path, template_path):
    today = date.today()
    try:
        from zoneinfo import ZoneInfo
        _now_mad = datetime.now(ZoneInfo("Europe/Madrid"))
        today_label = _now_mad.strftime("%Y-%m-%d %H:%M") + " (Madrid)"
    except Exception:
        today_label = today.isoformat()
    cutoffs = compute_cutoffs(today)
    opens = replay_opens(cache, cutoffs)

    present = set()
    for l, _ in cutoffs:
        for _, _, s in opens[l]:
            if s:
                present.add(s)
    states = [s for s in FUNNEL_ORDER if s in present] + sorted(s for s in present if s not in FUNNEL_ORDER)

    # Determinar cuántos cortes son fin-de-mes para clase CSS (cortes mensuales primero)
    monthly_count = 0
    for lbl, d in cutoffs:
        try:
            dnext = d + timedelta(days=1)
            if dnext.day == 1:
                monthly_count += 1
            else:
                break
        except Exception:
            break

    def hdr_cls(i):
        return "hdr-mes" if i < monthly_count else "hdr-dia"

    th_cells = "".join(
        f'<th class="{hdr_cls(i)}">{fmt_short(l)}</th>' for i, (l, _) in enumerate(cutoffs)
    )
    state_header = '<tr><th class="sticky-l1">Funnel</th><th class="sticky-l2">Estado</th>' + th_cells + '</tr>'
    chain_header = '<tr><th class="sticky-l1">Cadena</th>' + th_cells + '</tr>'

    taller_set_local = set(TALLER_STATUSES_ORDER)
    states_taller_idx = [i for i, s in enumerate(states) if s in taller_set_local]
    taller_first_idx = states_taller_idx[0] if states_taller_idx else -1
    taller_last_idx = states_taller_idx[-1] if states_taller_idx else -1

    def state_row(st):
        ft = FUNNEL_TAG.get(st, "?")
        color = FUNNEL_COLOR.get(ft, "#888")
        row_classes = []
        if st in taller_set_local:
            row_classes.append("row-taller")
            i = states.index(st)
            if i == taller_first_idx:
                row_classes.append("row-taller-first")
            if i == taller_last_idx:
                row_classes.append("row-taller-last")
        cls_attr = f' class="{" ".join(row_classes)}"' if row_classes else ""
        parts = [f'<tr{cls_attr}><td class="lbl-funnel" style="color:{color}">{ft}</td>',
                 f'<td class="lbl-name">{st}</td>']
        for l, _ in cutoffs:
            n = sum(1 for _, _, s in opens[l] if s == st)
            cls = "num n0" if n == 0 else "num"
            parts.append(f'<td class="{cls}">{n if n else "·"}</td>')
        parts.append('</tr>')
        return "".join(parts)

    # nuevas_por_chain_cutoff: {(ch, lbl): count_nuevas_esa_cadena_ese_corte}
    nuevas_por_chain_cutoff = {}
    for i, (lbl, d) in enumerate(cutoffs):
        if i < monthly_count:
            start_dt = datetime(d.year, d.month, 1, 0, 0, 0, tzinfo=timezone.utc)
        else:
            start_dt = datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=timezone.utc)
        end_dt = datetime(d.year, d.month, d.day, 23, 59, 59, tzinfo=timezone.utc)
        for key, t in cache.get("tickets", {}).items():
            created = parse_iso(t.get("created"))
            if not created or not (start_dt <= created <= end_dt):
                continue
            ch = chain_of(t.get("cliente", ""), t.get("loc", ""))
            nuevas_por_chain_cutoff[(ch, lbl)] = nuevas_por_chain_cutoff.get((ch, lbl), 0) + 1

    def chain_row(ch):
        cls_row = ' class="otros"' if ch == "Otros" else ''
        parts = [f'<tr{cls_row}><td class="lbl-name">{ch}</td>']
        for l, _ in cutoffs:
            n = sum(1 for _, x, _ in opens[l] if x == ch)
            new = nuevas_por_chain_cutoff.get((ch, l), 0)
            cls = "num n0" if n == 0 else "num"
            if new > 0:
                nuevas_txt = f'<span style="color:#16a34a;font-size:9px;font-weight:500;margin-left:3px" title="Nuevas creadas">(+{new})</span>'
                parts.append(f'<td class="{cls}">{n if n else chr(183)}{nuevas_txt}</td>')
            else:
                parts.append(f'<td class="{cls}">{n if n else chr(183)}</td>')
        parts.append('</tr>')
        return "".join(parts)

    state_rows = "\n".join(state_row(s) for s in states)
    chain_rows = "\n".join(chain_row(ch) for ch in CHAIN_ORDER)
    nuevas_total = {}
    for i, (lbl, d) in enumerate(cutoffs):
        if i < monthly_count:
            start_dt = datetime(d.year, d.month, 1, 0, 0, 0, tzinfo=timezone.utc)
        else:
            start_dt = datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=timezone.utc)
        end_dt = datetime(d.year, d.month, d.day, 23, 59, 59, tzinfo=timezone.utc)
        cnt = 0
        for key, t in cache.get("tickets", {}).items():
            created = parse_iso(t.get("created"))
            if not created:
                continue
            if start_dt <= created <= end_dt:
                cnt += 1
        nuevas_total[lbl] = cnt

    def _state_total_cell(l):
        n = len(opens[l])
        nv = nuevas_total.get(l, 0)
        if nv > 0:
            return f'<td class="num">{n} <span style="color:#16a34a;font-size:10px;font-weight:500" title="Nuevas creadas">(+{nv})</span></td>'
        return f'<td class="num">{n}</td>'
    state_total = ('<tr class="total"><td class="lbl-funnel"></td><td class="lbl-name">Total</td>'
                   + "".join(_state_total_cell(l) for l, _ in cutoffs) + '</tr>')
    chain_total = ('<tr class="total"><td class="lbl-name">Total</td>'
                   + "".join(f'<td class="num">{len(opens[l])}</td>' for l, _ in cutoffs) + '</tr>')

    today_total = len(opens[cutoffs[-1][0]])
    week_ago_total = len(opens[cutoffs[-6][0]]) if len(cutoffs) >= 6 else today_total
    dweek = today_total - week_ago_total
    arrow = "▲" if dweek > 0 else ("▼" if dweek < 0 else "→")
    dcolor = "#c0392b" if dweek > 0 else ("#1f8a4c" if dweek < 0 else "#7d8590")
    TALLER_STATUSES = set(TALLER_STATUSES_ORDER)
    en_taller = sum(
        1 for k, t in cache.get("tickets", {}).items()
        if t.get("current_status") in TALLER_STATUSES
    )

    estancadas_15 = sum(
        1 for k, t in cache.get("tickets", {}).items()
        if is_open(t.get("current_status"))
        and (days_since(last_status_change_dt(t)) or 0) > 15
    )

    # Pico = max histórico en taller (en lugar de max histórico de abiertas)
    peak = 0
    peak_lbl = ""
    for lbl, _ in cutoffs:
        n = sum(1 for _, _, s in opens[lbl] if s in TALLER_STATUSES)
        if n > peak:
            peak = n
            peak_lbl = lbl
    peak_pct = round(en_taller / peak * 100) if peak else 0
    peak_color = "#1f8a4c" if peak_pct <= 70 else ("#b8860b" if peak_pct <= 90 else "#c0392b")

    # Flechas in/out totales para el bloque "En taller"
    changes_today_kpi = compute_status_changes_today(cache)
    taller_in = sum(v["in"] for (st, _ch), v in changes_today_kpi.items() if st in TALLER_STATUSES)
    taller_out = sum(v["out"] for (st, _ch), v in changes_today_kpi.items() if st in TALLER_STATUSES)
    en_taller_flow_parts = []
    if taller_in > 0:
        en_taller_flow_parts.append(f'<span style="color:#c0392b;font-weight:500" title="entran hoy a taller">&#9650;{taller_in}</span>')
    if taller_out > 0:
        en_taller_flow_parts.append(f'<span style="color:#1f8a4c;font-weight:500" title="salen hoy de taller">&#9660;{taller_out}</span>')
    en_taller_flow = ('<div class="d" style="margin-top:2px;font-size:13px">'
                      + ' '.join(en_taller_flow_parts) + '</div>') if en_taller_flow_parts else ''

    # Cargar sustis cache para enriquecer Detalle y calcular pendientes críticos
    sustis_by_parent = {}
    try:
        from pathlib import Path as _PSusti
        for _cand in [_PSusti("cache") / "sustis_activas.json", out_path.parent.parent / "cache" / "sustis_activas.json"]:
            if _cand.exists():
                _su = json.loads(_cand.read_text(encoding="utf-8"))
                for _it in _su.get("items", []) or []:
                    _pk = _it.get("parent_key")
                    _ss = (_it.get("subtask_status") or "").strip()
                    if _pk and _ss:
                        sustis_by_parent[_pk] = _ss
                break
    except Exception:
        pass

    # KPI Pendientes a taller: estados antes del taller (no asignados aún a técnico)
    PRE_TALLER_STATUSES = {"Abierto", "Recepcionado SAT", "Pendiente recogida", "Gestionado transporte"}
    BLOQ_TRUE = {"sí", "si", "yes", "true", "1"}
    SUSTI_TRUE = {"sí", "si", "yes", "true", "1"}
    pendientes_taller = 0
    pendientes_taller_bloq = 0
    pendientes_taller_susti = 0
    pendientes_taller_critic = 0   # NUEVO: bloq=Sí AND susti en estado Solicitado
    pendientes_taller_keys = []
    for k, t in cache.get("tickets", {}).items():
        if t.get("current_status") not in PRE_TALLER_STATUSES:
            continue
        pendientes_taller += 1
        pendientes_taller_keys.append(k)
        is_bloq = (t.get("bloq") or "").strip().lower() in BLOQ_TRUE
        is_susti = (t.get("susti") or "").strip().lower() in SUSTI_TRUE
        if is_bloq:
            pendientes_taller_bloq += 1
        if is_susti:
            pendientes_taller_susti += 1
        ss = sustis_by_parent.get(k, "")
        if is_bloq and ss.lower() == "solicitado":
            pendientes_taller_critic += 1

    # KPI Gestión externa
    EXTERNA_STATUSES = {"Enviado a técnico externo", "Esperando respuesta cliente a presupuesto", "Pendiente definir servicio externo"}
    gestion_externa = 0
    gestion_externa_split = {"Carlos Mora": 0, "Sincronis": 0, "GEE": 0, "Innet": 0, "Otros": 0, "(Sin asignar)": 0}
    EXT_KEYS = {"carlos mora": "Carlos Mora", "carlos": "Carlos Mora",
                "sincronis": "Sincronis",
                "gee": "GEE",
                "innet": "Innet"}
    for k, t in cache.get("tickets", {}).items():
        if t.get("current_status") in EXTERNA_STATUSES:
            gestion_externa += 1
            tec_ext = (t.get("tec_externo") or "").strip().lower()
            matched = None
            for kw, lbl in EXT_KEYS.items():
                if kw in tec_ext:
                    matched = lbl; break
            if matched:
                gestion_externa_split[matched] += 1
            elif tec_ext:
                gestion_externa_split["Otros"] += 1
            else:
                gestion_externa_split["(Sin asignar)"] += 1

    # KPI Splits "En taller" — 4 subgrupos
    EN_TALLER_SPLIT_DEFS = [
        ("Sin presupuesto", "Pdte asignar / cola / preparación", {"Pendiente asignar técnico", "En cola taller", "En preparación presupuesto"}),
        ("Pdte enviar/confirmar ppto", "Presupuesto preparado / pdte confirmación", {"Presupuesto preparado pendiente de enviar", "Pendiente confirmación presupuesto"}),
        ("En reparación", "Esperando inicio + en reparación", {"Esperando inicio reparación", "En reparación"}),
        ("Inspección de salida", "Listo para salir del taller", {"Inspección de salida"}),
    ]
    en_taller_split = []
    for lbl, desc, states in EN_TALLER_SPLIT_DEFS:
        n = sum(1 for k, t in cache.get("tickets", {}).items()
                if t.get("current_status") in states)
        en_taller_split.append({"label": lbl, "desc": desc, "count": n,
                                "states": sorted(states)})

    # KPI Sustis solicitadas (pendientes de entregar) - del cache de sustis
    # KPI Salidas de taller hoy: transiciones hoy from=cualquier estado taller -> Devuelto/Resuelto/Finalizada
    SALIDA_DESTINOS = {"Devuelto a cliente", "Resuelto", "Finalizada"}
    salidas_taller = 0
    salidas_taller_tickets = set()
    _cutoff_hoy = datetime(today.year, today.month, today.day, 0, 0, 0, tzinfo=timezone.utc)
    for _kk, _t in cache.get("tickets", {}).items():
        for _tr in _t.get("transitions", []):
            if len(_tr) < 3: continue
            _ts, _fs, _to = _tr[0], _tr[1], _tr[2]
            if _fs not in TALLER_STATUSES: continue
            if _to not in SALIDA_DESTINOS: continue
            _dt = parse_iso(_ts)
            if not _dt or _dt < _cutoff_hoy: continue
            if _kk not in salidas_taller_tickets:
                salidas_taller_tickets.add(_kk)
                salidas_taller += 1

    # KPI Entradas a taller hoy: tickets que han cambiado HOY a uno de los 8 estados de taller
    # y siguen abiertos. Set para clic-filtrar.
    entradas_taller_keys = set()
    for _kk, _t in cache.get("tickets", {}).items():
        if not is_open(_t.get("current_status")):
            continue
        for _tr in _t.get("transitions", []):
            if len(_tr) < 3: continue
            _ts, _fs, _to = _tr[0], _tr[1], _tr[2]
            if _to not in TALLER_STATUSES: continue
            if _fs in TALLER_STATUSES: continue  # transición interna al taller no cuenta
            _dt = parse_iso(_ts)
            if not _dt or _dt < _cutoff_hoy: continue
            entradas_taller_keys.add(_kk)
            break
    entradas_taller_count = len(entradas_taller_keys)

    sustis_solicitadas = 0
    try:
        from pathlib import Path as _P2
        for cand in [_P2("cache") / "sustis_activas.json", out_path.parent.parent / "cache" / "sustis_activas.json"]:
            if cand.exists():
                _s = json.loads(cand.read_text(encoding="utf-8"))
                sustis_solicitadas = sum(1 for it in (_s.get("items") or [])
                                          if (it.get("subtask_status") or "").lower() in ("solicitado",))
                break
    except Exception:
        pass

    # Nuevas creadas por corte (mensuales = todo el mes; diarios = ese dia)
    # KPI Nuevas hoy: tickets creados hoy Y abiertos (cuadra con bucket azul de la tabla Abiertas hoy)
    nuevas_hoy_keys_set = set()
    for _kk, _t in cache.get("tickets", {}).items():
        if is_created_today(_t) and is_open(_t.get("current_status")):
            nuevas_hoy_keys_set.add(_kk)
    nuevas_hoy_count = len(nuevas_hoy_keys_set)

    # KPIs: Presupuestos hechos hoy y Equipos reparados hoy (= tickets únicos con to_status hoy)
    _cutoff_hoy_kpi = datetime(today.year, today.month, today.day, 0, 0, 0, tzinfo=timezone.utc)
    pres_hechos_hoy = set()
    eq_reparados_hoy = set()
    for _k, _t in cache.get("tickets", {}).items():
        for _tr in _t.get("transitions", []):
            if len(_tr) < 3: continue
            _ts, _fs, _to = _tr[0], _tr[1], _tr[2]
            _dt = parse_iso(_ts)
            if not _dt or _dt < _cutoff_hoy_kpi: continue
            if _to == "Presupuesto preparado pendiente de enviar":
                pres_hechos_hoy.add(_k)
            elif _to == "Inspección de salida":
                eq_reparados_hoy.add(_k)
    presupuestos_hechos_count = len(pres_hechos_hoy)
    equipos_reparados_count = len(eq_reparados_hoy)

    state_nuevas = ""  # Las nuevas se muestran en la fila Total entre paréntesis en verde

    chain_nuevas = ""  # ya no se usa (nuevas se muestran en cada celda chain_row)

    is_demo = "DEMO" in (cache.get("_meta", {}).get("note") or "")
    demo = ('<div class="banner">⚠ Vista previa con cache DEMO. El GitHub Action poblará el histórico completo.</div>'
            if is_demo else '')

    ah_header, ah_rows = build_abiertas_hoy_section(cache)
    avances_header, avances_rows, avances_detail_json, avances_extra = build_avances_section(cache)
    tec_header, tec_rows = build_tecnicos_section(cache, avances_extra)
    detalle_rows = build_detalle_section(cache, sustis_by_parent=sustis_by_parent)

    # Fase 2: pestaña "Por cadena"
    chains_html = ""
    try:
        from build_chains import build_chains_html
        airtable = None
        airtable_path = out_path.parent.parent / "cache" / "airtable_pedidos.json"
        if airtable_path.exists():
            airtable = json.loads(airtable_path.read_text(encoding="utf-8"))
        serial_year = {}
        sy_path = out_path.parent.parent / "data" / "serial_year.json"
        if sy_path.exists():
            serial_year = json.loads(sy_path.read_text(encoding="utf-8"))
        def _find_cache(name):
            from pathlib import Path as _P
            for cand in [_P("cache") / name, out_path.parent.parent / "cache" / name]:
                if cand.exists():
                    return json.loads(cand.read_text(encoding="utf-8"))
            return None
        sustis = _find_cache("sustis_activas.json")
        inmov = _find_cache("airtable_inmovilizado.json")
        chains_html = build_chains_html(cache, airtable, serial_year, sustis, inmov)
        from sustis_etl import build_sustis_global_html
        sustis_html = build_sustis_global_html(sustis, inmov)
    except Exception as e:
        chains_html = f'<p style="color:#c0392b;padding:20px">Error generando Fase 2: {e}</p>'
        sustis_html = chains_html

    # Track record: guardar snapshot de KPIs hoy en cache/kpi_history.json (antes del render)
    history = {}
    try:
        hist_path = None
        for _cand in [Path("cache") / "kpi_history.json", out_path.parent.parent / "cache" / "kpi_history.json"]:
            if _cand.parent.exists():
                hist_path = _cand
                break
        if hist_path is not None:
            if hist_path.exists():
                try:
                    history = json.loads(hist_path.read_text(encoding="utf-8"))
                except Exception:
                    history = {}
            history[today.isoformat()] = {
                "ts": today_label,
                "abiertas_hoy": today_total,
                "nuevas_hoy": nuevas_hoy_count,
                "presupuestos_hechos": presupuestos_hechos_count,
                "equipos_reparados": equipos_reparados_count,
                "entradas_taller_hoy": entradas_taller_count,
                "salidas_taller_hoy": salidas_taller,
                "pendientes_taller": pendientes_taller,
                "pendientes_taller_critic": pendientes_taller_critic,
                "en_taller": en_taller,
                "peak_pct": peak_pct,
                "gestion_externa": gestion_externa,
                "sustis_solicitadas": sustis_solicitadas,
                "estancadas_15d": estancadas_15,
            }
            hist_path.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"[kpi_history] no se pudo guardar: {e}")

    # Generar tabla Evolución KPIs
    EVOL_KPI_COLS = [
        ("abiertas_hoy", "Abiertas", "#0b3d91"),
        ("nuevas_hoy", "Nuevas hoy", "#3b82f6"),
        ("presupuestos_hechos", "Presupuestos hechos", "#0891b2"),
        ("equipos_reparados", "Equipos reparados", "#059669"),
        ("entradas_taller_hoy", "Entradas taller", "#1f6feb"),
        ("salidas_taller_hoy", "Salidas taller", "#16a34a"),
        ("pendientes_taller", "Pdtes. taller", "#a23b72"),
        ("en_taller", "En taller", "#1f6feb"),
        ("peak_pct", "% pico taller", "#475569"),
        ("gestion_externa", "Gest. externa", "#cb6f0a"),
        ("sustis_solicitadas", "Sustis solic.", "#7c3aed"),
        ("estancadas_15d", "Estancadas >15d", "#c0392b"),
    ]
    dates_sorted = sorted(history.keys())
    if dates_sorted:
        ek_hdr = '<tr><th class="sticky-l1" style="text-align:left;left:0;min-width:120px">Fecha</th>'
        for _kk, _lbl, _col in EVOL_KPI_COLS:
            ek_hdr += f'<th style="color:#fff">{_lbl}</th>'
        ek_hdr += '</tr>'
        ek_rows_html = []
        for _date in dates_sorted:
            _row = history.get(_date, {})
            cells = [f'<td class="lbl-name" style="left:0;min-width:120px"><b>{_date}</b></td>']
            for _kk, _lbl, _col in EVOL_KPI_COLS:
                _v = _row.get(_kk, "")
                if _kk == "peak_pct" and _v != "":
                    cells.append(f'<td class="num" style="color:{_col};font-weight:500">{_v}%</td>')
                elif _v != "" and _v != 0:
                    cells.append(f'<td class="num" style="color:{_col};font-weight:500">{_v}</td>')
                else:
                    cells.append('<td class="num n0">·</td>')
            ek_rows_html.append('<tr>' + "".join(cells) + '</tr>')
        evolucion_kpis = ek_hdr + "\n".join(ek_rows_html)
    else:
        evolucion_kpis = '<tr><th colspan="9" style="color:#94a3b8;font-style:italic">Aún no hay historial. Se irá llenando a partir de hoy.</th></tr>'

    # Año venta x tipo averia (tabla con filtros cadena+importe)
    try:
        anual_averias_html = build_anual_averias_section(cache)
    except Exception as _e:
        anual_averias_html = f'<div style="color:#c0392b;padding:14px">Error tabla averias: {_e}</div>'

    # Tiempos refactorizados (base 2026+1-ene abiertos + live)
    try:
        _tiempos = build_tiempos_section(cache, sustis_by_parent)
        tiempo_taller_html = _tiempos["taller"]
        tiempo_sustis_html = _tiempos["sustis"]
        taller_detail_html = _tiempos["detail"]
    except Exception as _e:
        import traceback; traceback.print_exc()
        tiempo_taller_html = f'<div style="color:#c0392b;padding:14px">Error tiempos: {_e}</div>'
        tiempo_sustis_html = tiempo_taller_html
        taller_detail_html = tiempo_taller_html

    # Tabla técnicos por día (mes a mes desde 1-ene-2026)
    try:
        tec_dia_html = build_tecnicos_dia_section(cache)
    except Exception as _e:
        tec_dia_html = f'<div style="color:#c0392b;padding:14px">Error tec-dia: {_e}</div>'

    # Mapa: cargar geocodes + construir markers de incidencias abiertas y sustis activas
    geo_entries = {}
    try:
        from pathlib import Path as _PG
        for _cand in [_PG("cache") / "geocodes.json", out_path.parent.parent / "cache" / "geocodes.json"]:
            if _cand.exists():
                geo_entries = json.loads(_cand.read_text(encoding="utf-8")).get("entries", {})
                break
    except Exception:
        pass

    def _addr_key_local(cliente, loc):
        import re as _re
        loc_c = _re.sub(r"\s+", " ", str(loc or "")).strip()
        cli_c = _re.sub(r"\s+", " ", str(cliente or "")).strip()
        if loc_c and len(loc_c) > 8:
            return loc_c.lower()
        if cli_c:
            return cli_c.lower()
        return ""

    map_markers_incidencias = []
    for _k, _t in cache.get("tickets", {}).items():
        if not is_open(_t.get("current_status")): continue
        _key = _addr_key_local(_t.get("cliente",""), _t.get("loc",""))
        _g = geo_entries.get(_key)
        if not _g or _g.get("lat") is None: continue
        _days = days_since(last_status_change_dt(_t))
        map_markers_incidencias.append({
            "lat": _g["lat"], "lon": _g["lon"],
            "key": _k,
            "cliente": (_t.get("cliente") or "")[:80],
            "loc": (_t.get("loc") or "")[:80],
            "status": _t.get("current_status", ""),
            "chain": chain_of(_t.get("cliente",""), _t.get("loc","")),
            "bloq": _t.get("bloq", ""),
            "susti": _t.get("susti", ""),
            "days": _days if _days is not None else 0,
            "susti_status": sustis_by_parent.get(_k, ""),
        })

    map_markers_sustis = []
    try:
        _sus_path = None
        from pathlib import Path as _PS
        for _cand in [_PS("cache") / "sustis_activas.json", out_path.parent.parent / "cache" / "sustis_activas.json"]:
            if _cand.exists():
                _sus_path = _cand; break
        if _sus_path:
            _sd = json.loads(_sus_path.read_text(encoding="utf-8"))
            _now = datetime.now(timezone.utc)
            for _it in _sd.get("items", []) or []:
                _key = _addr_key_local(_it.get("cliente","") or _it.get("parent_cliente",""), _it.get("loc",""))
                _g = geo_entries.get(_key)
                if not _g or _g.get("lat") is None: continue
                _fe = _it.get("fecha_envio")
                _dias = ""
                if _fe:
                    _fdt = parse_iso(_fe)
                    if _fdt:
                        if not _fdt.tzinfo: _fdt = _fdt.replace(tzinfo=timezone.utc)
                        _dias = (_now - _fdt).days
                map_markers_sustis.append({
                    "lat": _g["lat"], "lon": _g["lon"],
                    "key": _it.get("key",""),
                    "parent_key": _it.get("parent_key",""),
                    "cliente": (_it.get("cliente") or "")[:80],
                    "loc": (_it.get("loc") or "")[:80],
                    "subtask_status": _it.get("subtask_status",""),
                    "consola": _it.get("consola_susti",""),
                    "manipulo": _it.get("manipulo_susti",""),
                    "dias": _dias if _dias != "" else 0,
                })
    except Exception:
        pass

    html = template_path.read_text(encoding="utf-8")
    repl = {
        "__TODAY__": today_label,
        "__DEMO_BANNER__": demo,
        "__TODAY_TOTAL__": str(today_total),
        "__DELTA_COLOR__": dcolor,
        "__DELTA_ARROW__": arrow,
        "__DELTA_ABS__": str(abs(dweek)),
        "__ESTANCADAS_15__": str(estancadas_15),
        "__PEAK__": str(peak),
        "__PEAK_LBL__": peak_lbl,
        "__PEAK_PCT__": str(peak_pct),
        "__PEAK_COLOR__": peak_color,
        "__EN_TALLER__": str(en_taller),
        "__EN_TALLER_FLOW__": en_taller_flow,
        "__AH_HEADER__": ah_header,
        "__AH_ROWS__": ah_rows,
        "__TEC_HEADER__": tec_header,
        "__TEC_ROWS__": tec_rows,
        "__AVANCES_HEADER__": avances_header,
        "__AVANCES_ROWS__": avances_rows,
        "__AVANCES_DETAIL_JSON__": avances_detail_json,
        "__DETALLE_ROWS__": detalle_rows,
        "__STATE_HEADER__": state_header,
        "__STATE_ROWS__": state_rows,
        "__STATE_TOTAL__": state_total,
        "__STATE_NUEVAS__": state_nuevas,
        "__CHAIN_HEADER__": chain_header,
        "__CHAIN_ROWS__": chain_rows,
        "__CHAIN_TOTAL__": chain_total,
        "__CHAIN_NUEVAS__": chain_nuevas,
        "__PENDIENTES_TALLER__": str(pendientes_taller),
        "__PENDIENTES_TALLER_BLOQ__": str(pendientes_taller_bloq),
        "__GESTION_EXTERNA__": str(gestion_externa),
        "__GESTION_EXTERNA_SPLIT__": json.dumps(gestion_externa_split),
        "__EN_TALLER_SPLIT__": json.dumps(en_taller_split),
        "__SUSTIS_SOLICITADAS__": str(sustis_solicitadas),
        "__SALIDAS_TALLER__": str(salidas_taller),
        "__NUEVAS_HOY__": str(nuevas_hoy_count),
        "__PRES_HECHOS__": str(presupuestos_hechos_count),
        "__EQ_REPARADOS__": str(equipos_reparados_count),
        "__EVOLUCION_KPIS__": evolucion_kpis,
        "__ANUAL_AVERIAS__": anual_averias_html,
        "__TIEMPO_TALLER__": tiempo_taller_html,
        "__TEC_DIA_SECTION__": tec_dia_html,
        "__TIEMPO_SUSTIS__": tiempo_sustis_html,
        "__TALLER_DETAIL__": taller_detail_html,
        "__MAP_INCIDENCIAS__": json.dumps(map_markers_incidencias),
        "__MAP_SUSTIS__": json.dumps(map_markers_sustis),
        "__CHAINS_HTML__": chains_html,
        "__SUSTIS_HTML__": sustis_html,
        "__PENDIENTES_TALLER_SUSTI__": str(pendientes_taller_susti),
        "__ENTRADAS_TALLER__": str(entradas_taller_count),
        "__KPI_PENDIENTES_TALLER_KEYS__": json.dumps(sorted(pendientes_taller_keys)),
        "__KPI_PRES_HECHOS_KEYS__": json.dumps(sorted(pres_hechos_hoy)),
        "__KPI_EQ_REPARADOS_KEYS__": json.dumps(sorted(eq_reparados_hoy)),
        "__KPI_ENTRADAS_TALLER_KEYS__": json.dumps(sorted(entradas_taller_keys)),
        "__KPI_SALIDAS_TALLER_KEYS__": json.dumps(sorted(salidas_taller_tickets)),
        "__KPI_NUEVAS_HOY_KEYS__": json.dumps(sorted(nuevas_hoy_keys_set)),
        "__PENDIENTES_TALLER_CRITIC__": str(pendientes_taller_critic),
    }
    for k, v in repl.items():
        html = html.replace(k, v)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cache", required=True)
    ap.add_argument("--output-xlsx", default=None)
    ap.add_argument("--output-html", default=None)
    ap.add_argument("--template", default="scripts/template.html")
    args = ap.parse_args()

    cache = json.loads(Path(args.cache).read_text(encoding="utf-8"))
    if args.output_xlsx:
        build_excel(cache, Path(args.output_xlsx))
        print(f"Wrote {args.output_xlsx}")
    if args.output_html:
        build_html(cache, Path(args.output_html), Path(args.template))
        print(f"Wrote {args.output_html}")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
